import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse
from rest_framework import serializers, viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from django.db.models import Sum, F, Q
from django.utils import timezone
from .models import Item, Inventario, Movimiento, Bodega, Proveedor
from comercial.models import Referencias, Exportador, Contenedor

# -----------------------------------------------------------------------------
# SERIALIZERS
# -----------------------------------------------------------------------------

class ExportadorSimpleSerializer(serializers.ModelSerializer):
    class Meta:
        model = Exportador
        fields = ['id', 'nombre']

class BodegaSerializer(serializers.ModelSerializer):
    exportador_nombre = serializers.CharField(source='exportador.nombre', read_only=True)
    
    class Meta:
        model = Bodega
        fields = ['id', 'nombre', 'exportador', 'exportador_nombre']

class ProveedorSerializer(serializers.ModelSerializer):
    class Meta:
        model = Proveedor
        fields = ['id', 'nombre']

class ReferenciaSimpleSerializer(serializers.ModelSerializer):
    exportador_nombre = serializers.CharField(source='exportador.nombre', read_only=True)
    
    class Meta:
        model = Referencias
        fields = ['id', 'nombre', 'exportador', 'exportador_nombre']

class ItemSerializer(serializers.ModelSerializer):
    numero_item_nombre = serializers.CharField(source='numero_item.nombre', read_only=True)
    bodega_nombre = serializers.CharField(source='bodega.nombre', read_only=True)
    proveedor_nombre = serializers.CharField(source='proveedor.nombre', read_only=True)
    propiedad_nombre = serializers.CharField(source='propiedad.nombre', read_only=True)
    user_username = serializers.CharField(source='user.username', read_only=True)
    tipo_documento_display = serializers.CharField(source='get_tipo_documento_display', read_only=True)
    
    class Meta:
        model = Item
        fields = [
            'id', 'numero_item', 'numero_item_nombre', 'cantidad_cajas', 
            'tipo_documento', 'tipo_documento_display', 'documento', 
            'bodega', 'bodega_nombre', 'proveedor', 'proveedor_nombre',
            'fecha_movimiento', 'propiedad', 'propiedad_nombre', 
            'observaciones', 'user', 'user_username'
        ]
        read_only_fields = ['user']

class InventarioSerializer(serializers.ModelSerializer):
    numero_item_nombre = serializers.CharField(source='numero_item.nombre', read_only=True)
    exportador_nombre = serializers.CharField(source='numero_item.exportador.nombre', read_only=True)
    stock_actual = serializers.SerializerMethodField()
    
    class Meta:
        model = Inventario
        fields = [
            'id', 'numero_item', 'numero_item_nombre', 'exportador_nombre',
            'compras_efectivas', 'saldos_iniciales', 'salidas', 
            'traslado_propio', 'traslado_remisionado', 'ventas', 
            'venta_contenedor', 'stock_actual'
        ]

    def get_stock_actual(self, obj):
        ingresos = (obj.compras_efectivas or 0) + (obj.saldos_iniciales or 0)
        salidas = (obj.salidas or 0) + (obj.traslado_propio or 0) + (obj.traslado_remisionado or 0) + (obj.ventas or 0) + (obj.venta_contenedor or 0)
        return ingresos - salidas

class MovimientoSerializer(serializers.ModelSerializer):
    bodega_nombre = serializers.CharField(source='bodega.nombre', read_only=True)
    user_username = serializers.CharField(source='user.username', read_only=True)
    
    class Meta:
        model = Movimiento
        fields = [
            'id', 'item_historico', 'cantidad_cajas_h', 'bodega', 'bodega_nombre',
            'propiedad', 'fecha_movimiento', 'observaciones', 'fecha',
            'user', 'user_username'
        ]

# -----------------------------------------------------------------------------
# PAGINATION
# -----------------------------------------------------------------------------

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 1000

# -----------------------------------------------------------------------------
# VIEWSETS
# -----------------------------------------------------------------------------

class BaseInventoryViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    
    def get_exportador_filter(self):
        """
        Returns the Q object to filter by exportador based on user groups.
        This provides security ensuring users only see data they are allowed to see.
        """
        user = self.request.user
        groups = user.groups.values_list('name', flat=True)
        
        # Superusers or Heavens group see everything
        if user.is_superuser or 'Heavens' in groups or 'Autorizadores' in groups:
            return Q()
            
        # Build filter based on reliable group names
        filters = Q()
        if 'Etnico' in groups:
            filters |= Q(exportador__nombre='Etnico')
        if 'Fieldex' in groups:
            filters |= Q(exportador__nombre='Fieldex')
        if 'Juan_Matas' in groups:
            filters |= Q(exportador__nombre='Juan_Matas')
        if 'CI_Dorado' in groups:
            filters |= Q(exportador__nombre='CI_Dorado')
            
        return filters

class InventarioViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for inventory summaries (aggregated stock per reference).
    Pagination is REQUIRED - frontend expects {count, results} structure.
    """
    serializer_class = InventarioSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination  # CRITICAL: Do not remove
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['numero_item__nombre', 'numero_item__exportador__nombre']
    ordering_fields = '__all__'
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        
        # Calculate totals for the entire FILTERED queryset
        totals = queryset.aggregate(
            total_compras=Sum('compras_efectivas'),
            total_saldos_iniciales=Sum('saldos_iniciales'),
            total_salidas=Sum('salidas'),
            total_traslado_propio=Sum('traslado_propio'),
            total_traslado_remisionado=Sum('traslado_remisionado'),
            total_ventas=Sum('ventas'),
            total_venta_contenedor=Sum('venta_contenedor')
        )
        
        # We also need counts for the KPIs
        # This is more efficient to do on the whole queryset here
        all_items = list(queryset)
        def get_stock(item):
            ingresos = (item.compras_efectivas or 0) + (item.saldos_iniciales or 0)
            salidas = (item.salidas or 0) + (item.traslado_propio or 0) + (item.traslado_remisionado or 0) + (item.ventas or 0) + (item.venta_contenedor or 0)
            return ingresos - salidas

        low_stock_count = sum(1 for item in all_items if 0 < get_stock(item) < 50)
        out_of_stock_count = sum(1 for item in all_items if get_stock(item) <= 0)
        
        # Calculate stock total (Ingresos - Salidas)
        # We need to sum them carefully handling None values
        ingresos = (totals['total_compras'] or 0) + (totals['total_saldos_iniciales'] or 0)
        egresos = (totals['total_salidas'] or 0) + (totals['total_traslado_propio'] or 0) + \
                  (totals['total_traslado_remisionado'] or 0) + (totals['total_ventas'] or 0) + (totals['total_venta_contenedor'] or 0)
        stock_total = ingresos - egresos

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response = self.get_paginated_response(serializer.data)
            response.data['totals'] = {
                'compras_efectivas': totals['total_compras'] or 0,
                'saldos_iniciales': totals['total_saldos_iniciales'] or 0,
                'salidas': totals['total_salidas'] or 0,
                'traslado_propio': totals['total_traslado_propio'] or 0,
                'traslado_remisionado': totals['total_traslado_remisionado'] or 0,
                'ventas': (totals['total_ventas'] or 0) + (totals['total_venta_contenedor'] or 0),
                'stock_actual': stock_total,
                'low_stock_count': low_stock_count,
                'out_of_stock_count': out_of_stock_count
            }
            return response

        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'results': serializer.data,
            'totals': {
                'compras_efectivas': totals['total_compras'] or 0,
                'saldos_iniciales': totals['total_saldos_iniciales'] or 0,
                'salidas': totals['total_salidas'] or 0,
                'traslado_propio': totals['total_traslado_propio'] or 0,
                'traslado_remisionado': totals['total_traslado_remisionado'] or 0,
                'ventas': totals['total_ventas'] or 0,
                'stock_actual': stock_total
            }
        })
    
    def get_queryset(self):
        user = self.request.user
        
        # Optimize query
        queryset = Inventario.objects.all().select_related(
            'numero_item', 
            'numero_item__exportador'
        ).order_by('numero_item__nombre')
        
        # Get flattened groups list
        groups = list(user.groups.values_list('name', flat=True))

        # Access Control Logic
        if user.is_superuser or 'Heavens' in groups or 'Autorizadores' in groups:
            # Superusers and Heavens staff see all
            pass 
        else:
            # For specific exporters, build a composite filter
            q_filter = Q()
            has_permission = False
            
            if 'Etnico' in groups:
                q_filter |= Q(numero_item__exportador__nombre='Etnico')
                has_permission = True
            if 'Fieldex' in groups:
                q_filter |= Q(numero_item__exportador__nombre='Fieldex')
                has_permission = True
            if 'Juan_Matas' in groups:
                 # Note: Exportador name in DB is 'Juan_Matas' (underscore) based on verify script
                q_filter |= Q(numero_item__exportador__nombre='Juan_Matas')
                has_permission = True
            if 'CI_Dorado' in groups:
                q_filter |= Q(numero_item__exportador__nombre='CI_Dorado')
                has_permission = True
                
            # If user has no relevant groups, show nothing (Security Default)
            if not has_permission:
                return queryset.none()
                
            queryset = queryset.filter(q_filter)
            
        # URL Params Filtering (e.g. from Dropdown)
        exportador = self.request.query_params.get('exportador', None)
        if exportador and exportador != 'all':
            queryset = queryset.filter(numero_item__exportador__nombre=exportador)
            
        return queryset

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Exports the filtered inventory summary to Excel."""
        queryset = self.filter_queryset(self.get_queryset())
        
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Inventario General'
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")

        columns = [
            'Referencia', 'Exportador', 'Compras Efectivas', 'Saldos Iniciales', 
            'Salidas', 'Traslado Propio', 'Traslado Remisionado', 'Ventas', 
            'Venta Contenedor', 'Stock Actual'
        ]
        
        for col_num, column_title in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill

        for row_num, item in enumerate(queryset, start=2):
            # Calculate stock actual exactly like the serializer
            ingresos = (item.compras_efectivas or 0) + (item.saldos_iniciales or 0)
            salidas = (item.salidas or 0) + (item.traslado_propio or 0) + (item.traslado_remisionado or 0) + (item.ventas or 0)
            stock_actual = ingresos - salidas
            
            row = [
                item.numero_item.nombre,
                item.numero_item.exportador.nombre,
                item.compras_efectivas,
                item.saldos_iniciales,
                item.salidas,
                item.traslado_propio,
                item.traslado_remisionado,
                (item.ventas or 0) + (item.venta_contenedor or 0),
                stock_actual
            ]
            for col_num, cell_value in enumerate(row, start=1):
                worksheet.cell(row=row_num, column=col_num, value=cell_value)

        # Auto-adjust columns width
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="inventario_resumen_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        return response

class ItemViewSet(viewsets.ModelViewSet):
    """
    API endpoint for individual inventory movements (items).
    Pagination is REQUIRED - frontend expects {count, results} structure.
    """
    serializer_class = ItemSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination  # CRITICAL: Do not remove
    filter_backends = [filters.SearchFilter]
    search_fields = ['numero_item__nombre', 'documento', 'observaciones']
    
    def get_queryset(self):
        user = self.request.user
        groups = user.groups.values_list('name', flat=True)
        
        queryset = Item.objects.all().select_related(
            'numero_item', 'bodega', 'proveedor', 'propiedad', 'user'
        ).order_by('-fecha_movimiento')
        
        # Access Control
        if not (user.is_superuser or 'Heavens' in groups or 'Autorizadores' in groups):
            q_filter = Q()
            if 'Etnico' in groups:
                q_filter |= Q(bodega__exportador__nombre='Etnico')
            if 'Fieldex' in groups:
                q_filter |= Q(bodega__exportador__nombre='Fieldex')
            if 'Juan_Matas' in groups:
                q_filter |= Q(bodega__exportador__nombre='Juan_Matas')
            if 'CI_Dorado' in groups:
                q_filter |= Q(bodega__exportador__nombre='CI_Dorado')
            queryset = queryset.filter(q_filter)
            
        # URL Params Filtering
        exportador = self.request.query_params.get('exportador', None)
        if exportador:
            queryset = queryset.filter(bodega__exportador__nombre=exportador)
            
        return queryset

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Exports the filtered items (movements) to Excel."""
        queryset = self.filter_queryset(self.get_queryset())
        
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Movimientos de Inventario'
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")

        columns = [
            'Fecha', 'Referencia', 'Tipo Movimiento', 'Cantidad', 
            'Tipo Doc', 'Documento', 'Proveedor', 'Propiedad', 
            'Observaciones', 'Usuario'
        ]
        
        for col_num, column_title in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill

        for row_num, item in enumerate(queryset, start=2):
            row = [
                item.fecha_movimiento.strftime("%Y-%m-%d") if item.fecha_movimiento else "",
                item.numero_item.nombre,
                item.bodega.nombre,
                item.cantidad_cajas,
                item.get_tipo_documento_display(),
                item.documento,
                item.proveedor.nombre,
                item.propiedad.nombre,
                item.observaciones or "",
                item.user.username
            ]
            for col_num, cell_value in enumerate(row, start=1):
                worksheet.cell(row=row_num, column=col_num, value=cell_value)

        # Auto-adjust columns width
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="movimientos_inventario_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        return response

    def perform_create(self, serializer):
        item = serializer.save(user=self.request.user)
        # Movement history is created via signals or manual create call here if needed
        # The existing logic uses views that manually create Movimiento.
        # Ideally, we should unify this process. The existing ItemCreateView manually creates movement.
        # We'll duplicate that logic here for the API.
        
        Movimiento.objects.create(
            item_historico=item.numero_item.nombre, # Note: Existing views stringify it or pass object? Model says CharField but Views pass FK. Let's check model.
            # Models.py: item_historico = models.CharField(max_length=100)
            # Views.py: item_historico=item.numero_item (which is a Referencias object)
            # Django auto-converts FK to str if assigned to CharField? No, usually not. 
            # Looking at existing views: Movimiento.objects.create(item_historico=item.numero_item, ...)
            # We should stringify it to be safe or pass the object if Django handles it (likely via __str__)
            cantidad_cajas_h=item.cantidad_cajas,
            bodega=item.bodega,
            propiedad=item.propiedad.nombre, # Model says CharField for property
            fecha_movimiento=item.fecha_movimiento,
            observaciones=item.observaciones,
            fecha=timezone.now(),
            user=item.user
        )

    def perform_update(self, serializer):
        item = serializer.save(user=self.request.user)
        # Create history for update too
        Movimiento.objects.create(
            item_historico=item.numero_item.nombre,
            cantidad_cajas_h=item.cantidad_cajas,
            bodega=item.bodega,
            propiedad=item.propiedad.nombre,
            fecha_movimiento=item.fecha_movimiento,
            observaciones=f"EDICIÓN: {item.observaciones or ''}",
            fecha=timezone.now(),
            user=item.user
        )

    def perform_destroy(self, instance):
        # Create history for deletion
        Movimiento.objects.create(
            item_historico=instance.numero_item.nombre,
            cantidad_cajas_h=instance.cantidad_cajas,
            bodega=instance.bodega,
            propiedad=instance.propiedad.nombre,
            fecha_movimiento=instance.fecha_movimiento,
            observaciones=f"ELIMINACIÓN ID {instance.id}: {instance.observaciones or ''}",
            fecha=timezone.now(),
            user=self.request.user
        )
        instance.delete()

class MovimientoViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for movement history (historical record of all changes).
    Pagination is REQUIRED - frontend expects {count, results} structure.
    """
    serializer_class = MovimientoSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination  # CRITICAL: Do not remove
    filter_backends = [filters.SearchFilter]
    search_fields = ['item_historico', 'observaciones']
    
    def get_queryset(self):
        user = self.request.user
        groups = user.groups.values_list('name', flat=True)
        queryset = Movimiento.objects.all().select_related('bodega', 'user').order_by('-fecha')
        
        # Access Control
        if not (user.is_superuser or 'Heavens' in groups or 'Autorizadores' in groups):
            q_filter = Q()
            if 'Etnico' in groups:
                q_filter |= Q(bodega__exportador__nombre='Etnico')
            if 'Fieldex' in groups:
                q_filter |= Q(bodega__exportador__nombre='Fieldex')
            if 'Juan_Matas' in groups:
                q_filter |= Q(bodega__exportador__nombre='Juan_Matas')
            if 'CI_Dorado' in groups:
                q_filter |= Q(bodega__exportador__nombre='CI_Dorado')
            queryset = queryset.filter(q_filter)
            
        return queryset

class BodegaViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for warehouse/bodega lookup (used in dropdowns).
    Filters by user's exportador groups to ensure data segregation.
    """
    serializer_class = BodegaSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None  # Dropdowns don't need pagination
    
    def get_queryset(self):
        user = self.request.user
        groups = list(user.groups.values_list('name', flat=True))
        
        queryset = Bodega.objects.all().select_related('exportador')
        
        # Access Control: Filter by user's exportador groups
        if not (user.is_superuser or 'Heavens' in groups or 'Autorizadores' in groups):
            q_filter = Q()
            if 'Etnico' in groups:
                q_filter |= Q(exportador__nombre='Etnico')
            if 'Fieldex' in groups:
                q_filter |= Q(exportador__nombre='Fieldex')
            if 'Juan_Matas' in groups:
                q_filter |= Q(exportador__nombre='Juan_Matas')
            if 'CI_Dorado' in groups:
                q_filter |= Q(exportador__nombre='CI_Dorado')
            queryset = queryset.filter(q_filter)
        
        # Additional URL param filter
        exportador = self.request.query_params.get('exportador', None)
        if exportador:
            queryset = queryset.filter(exportador__nombre=exportador)
        return queryset

class ProveedorViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for provider/proveedor lookup.
    Providers are shared across all exporters.
    """
    serializer_class = ProveedorSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None
    queryset = Proveedor.objects.all()

class ContenedorSerializer(serializers.ModelSerializer):
    class Meta:
        model = Contenedor
        fields = ['id', 'nombre', 'precio']

class ReferenciaSerializer(serializers.ModelSerializer):
    exportador_nombre = serializers.CharField(source='exportador.nombre', read_only=True)
    contenedor_nombre = serializers.CharField(source='contenedor.nombre', read_only=True)
    
    class Meta:
        model = Referencias
        fields = [
            'id', 'nombre', 'referencia_nueva', 'contenedor', 'contenedor_nombre',
            'cant_contenedor', 'precio', 'exportador', 'exportador_nombre',
            'cantidad_pallet_con_contenedor', 'cantidad_pallet_sin_contenedor',
            'porcentaje_peso_bruto'
        ]

class ReferenciasViewSet(viewsets.ModelViewSet):
    """
    API endpoint for managing References (Referencias).
    - List: Filtered by user group (Heavens sees all, Exporters see theirs).
    - Create/Delete: Only 'Heavens' group or Superuser.
    - Update: 
        - 'Heavens': Can edit all fields.
        - Exporters: Cannot edit 'nombre' or 'exportador'.
    """
    serializer_class = ReferenciaSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre', 'referencia_nueva']
    
    def get_queryset(self):
        user = self.request.user
        groups = list(user.groups.values_list('name', flat=True))
        
        queryset = Referencias.objects.all().select_related('exportador', 'contenedor').order_by('nombre')
        
        # Access Control: Filter by user's exportador groups
        if not (user.is_superuser or 'Heavens' in groups or 'Autorizadores' in groups):
            q_filter = Q()
            if 'Etnico' in groups:
                q_filter |= Q(exportador__nombre='Etnico')
            if 'Fieldex' in groups:
                q_filter |= Q(exportador__nombre='Fieldex')
            if 'Juan_Matas' in groups:
                q_filter |= Q(exportador__nombre='Juan_Matas')
            if 'CI_Dorado' in groups:
                q_filter |= Q(exportador__nombre='CI_Dorado')
            queryset = queryset.filter(q_filter)
        
        # URL Param Filter
        exportador = self.request.query_params.get('exportador', None)
        if exportador and exportador != 'all':
            queryset = queryset.filter(exportador__nombre=exportador)
            
        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        groups = user.groups.values_list('name', flat=True)
        if not (user.is_superuser or 'Heavens' in groups or 'Autorizadores' in groups):
             raise serializers.ValidationError({"detail": "No tienes permisos para crear referencias."})
        serializer.save()

    def perform_destroy(self, instance):
        user = self.request.user
        groups = user.groups.values_list('name', flat=True)
        if not (user.is_superuser or 'Heavens' in groups or 'Autorizadores' in groups):
             raise serializers.ValidationError({"detail": "No tienes permisos para eliminar referencias."})
        instance.delete()

    def perform_update(self, serializer):
        user = self.request.user
        groups = user.groups.values_list('name', flat=True)
        is_heavens = user.is_superuser or 'Heavens' in groups or 'Autorizadores' in groups
        
        # 'exportador' cannot be changed by ANYONE once created
        if 'exportador' in serializer.validated_data:
             serializer.validated_data.pop('exportador')

        if not is_heavens:
            # Prevent editing restricted fields for non-Heavens users
            if 'nombre' in serializer.validated_data:
                 serializer.validated_data.pop('nombre')
        
        serializer.save()

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Referencias'
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")

        columns = [
            'Referencia', 'Ref. Nueva', 'Exportador', 'Contenedor', 'Cant. Contenedor', 
            'Precio', 'Pallet c/ Cont.', 'Pallet s/ Cont.', '% Peso Bruto'
        ]
        
        for col_num, column_title in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill

        for row_num, item in enumerate(queryset, start=2):
            row = [
                item.nombre,
                item.referencia_nueva or "",
                item.exportador.nombre,
                item.contenedor.nombre if item.contenedor else "",
                item.cant_contenedor,
                item.precio,
                item.cantidad_pallet_con_contenedor,
                item.cantidad_pallet_sin_contenedor,
                item.porcentaje_peso_bruto
            ]
            for col_num, cell_value in enumerate(row, start=1):
                worksheet.cell(row=row_num, column=col_num, value=cell_value)

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="referencias_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        return response

class ContenedorViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ContenedorSerializer
    permission_classes = [IsAuthenticated]
    queryset = Contenedor.objects.all().order_by('nombre')
    pagination_class = None

import io
from django.contrib.auth.decorators import user_passes_test, login_required
from django.http import HttpResponse
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from .models import DetallePedido


# -----------Funcion para permisos por grupo ---------------------
def es_miembro_del_grupo(nombre_grupo):
    def es_miembro(user):
        return user.groups.filter(name=nombre_grupo).exists()

    return es_miembro

@login_required
@user_passes_test(es_miembro_del_grupo('Heavens'), login_url='home')
def exportar_detalles_pedidos_excel(request):
    # Crear un buffer en memoria
    output = io.BytesIO()

    # Crear un libro de trabajo (no usar write_only para poder modificar propiedades)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Detalles De Pedidos General'

    # Definir estilos para los encabezados
    header_font = Font(name='Arial', bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Estilos para filas alternas
    even_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Estilos para totales
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # Definir las columnas
    columnas = [
        'Pedido', 'F Entrega', 'Exportador', 'AWB', 'Cliente', 'Fruta', 'Presentacion', 'Cajas Solicitadas',
        'Peso Presentacion', 'kilos', 'Cajas Enviadas', 'Kilos Enviados', 'Diferencia', 'Tipo Caja',
        'Referencia', 'Stiker', 'Lleva Contenedor', 'Ref Contenedor', 'Cant Contenedor', 'Tarifa utilidad', 'Tarifa Recuperacion',
        'Valor x Caja USD', 'Valor X Producto', 'No Cajas NC', 'Valor NC', 'Afecta utilidad',
        'Valor Total utilidad Producto','Total Recuperacion X Producto',  'Precio Proforma', 'Observaciones'
    ]
    
    # Anchos aproximados para cada columna
    column_widths = {
        'Pedido': 10, 'F Entrega': 12, 'Exportador': 20, 'AWB': 15, 'Cliente': 20,
        'Fruta': 15, 'Presentacion': 15, 'Cajas Solicitadas': 15, 'Peso Presentacion': 15,
        'kilos': 10, 'Cajas Enviadas': 15, 'Kilos Enviados': 15, 'Diferencia': 12,
        'Tipo Caja': 15, 'Referencia': 15, 'Stiker': 15, 'Lleva Contenedor': 15,
        'Ref Contenedor': 15, 'Cant Contenedor': 15, 'Tarifa utilidad': 15,
        'Tarifa Recuperacion': 15, 'Valor x Caja USD': 15, 'Valor X Producto': 15,
        'No Cajas NC': 12, 'Valor NC': 12, 'Afecta utilidad': 15,
        'Valor Total utilidad Producto': 22, 'Total Recuperacion X Producto': 22,
        'Precio Proforma': 15, 'Observaciones': 30
    }
    
    # Identificar columnas numéricas para aplicar formato después
    numeric_columns = [
        'Cajas Solicitadas', 'Peso Presentacion', 'kilos', 'Cajas Enviadas', 'Kilos Enviados',
        'Diferencia', 'Cant Contenedor', 'Tarifa utilidad', 'Tarifa Recuperacion',
        'Valor x Caja USD', 'Valor X Producto', 'No Cajas NC', 'Valor NC',
        'Valor Total utilidad Producto', 'Total Recuperacion X Producto', 'Precio Proforma'
    ]
    numeric_indices = [columnas.index(col) for col in numeric_columns]
    
    # Columnas monetarias para formato de moneda
    currency_columns = [
        'Tarifa utilidad', 'Tarifa Recuperacion', 'Valor x Caja USD', 'Valor X Producto', 
        'Valor NC', 'Valor Total utilidad Producto', 'Total Recuperacion X Producto', 'Precio Proforma'
    ]
    currency_indices = [columnas.index(col) for col in currency_columns]

    # Agregar encabezados con estilos
    for col_idx, column_title in enumerate(columnas, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Ajustar ancho de columnas
        worksheet.column_dimensions[get_column_letter(col_idx)].width = column_widths.get(column_title, 15)

    # Obtener filtros desde el request
    numero_pedido_inicial = request.POST.get('numero_pedido_inicial')
    numero_pedido_final = request.POST.get('numero_pedido_final')

    try:
        if numero_pedido_inicial and numero_pedido_final:
            numero_pedido_inicial = int(numero_pedido_inicial)
            numero_pedido_final = int(numero_pedido_final)
            queryset = DetallePedido.objects.select_related(
                'pedido__exportadora',
                'pedido__cliente',
                'fruta',
                'presentacion',
                'tipo_caja',
                'referencia',
                # Agrega otros campos relacionados según sea necesario
            ).filter(pedido__pk__range=(numero_pedido_inicial, numero_pedido_final))
        else:
            queryset = DetallePedido.objects.select_related(
                'pedido__exportadora',
                'pedido__cliente',
                'fruta',
                'presentacion',
                'tipo_caja',
                'referencia',
                # Agrega otros campos relacionados según sea necesario
            ).all()
        
        # Variables para totales
        totals = {idx: 0 for idx in numeric_indices}
        row_idx = 2  # Comenzamos en la fila 2 después del encabezado

        # Iterar sobre el queryset sin almacenar en caché
        for detalle in queryset.iterator():
            fila = [
                detalle.pedido.pk,
                detalle.pedido.fecha_entrega,
                detalle.pedido.exportadora.nombre,
                detalle.pedido.awb,
                detalle.pedido.cliente.nombre,
                detalle.fruta.nombre,
                detalle.presentacion.nombre,
                detalle.cajas_solicitadas,
                detalle.presentacion_peso,
                detalle.kilos,
                detalle.cajas_enviadas,
                detalle.kilos_enviados,
                detalle.diferencia,
                detalle.tipo_caja.nombre,
                detalle.referencia.nombre,
                detalle.stickers,
                detalle.lleva_contenedor,
                detalle.referencia_contenedor,
                detalle.cantidad_contenedores,
                detalle.tarifa_utilidad,
                detalle.tarifa_recuperacion,
                detalle.valor_x_caja_usd,
                detalle.valor_x_producto,
                detalle.no_cajas_nc,
                detalle.valor_nota_credito_usd,
                detalle.afecta_utilidad,
                detalle.valor_total_utilidad_x_producto,
                detalle.valor_total_recuperacion_x_producto,
                detalle.precio_proforma,
                detalle.observaciones,
            ]
            
            # Sumar a los totales
            for idx in numeric_indices:
                if isinstance(fila[idx], (int, float)) and fila[idx] is not None:
                    totals[idx] += fila[idx]

            # Aplicar estilo a filas alternas
            for col_idx, value in enumerate(fila, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Side(style='thin')
                
                # Formato de fecha
                if col_idx == 2 and isinstance(value, datetime):  # F Entrega
                    cell.number_format = 'DD/MM/YYYY'
                
                # Formato de números
                if col_idx - 1 in numeric_indices:
                    cell.number_format = '#,##0.00'
                
                # Formato de moneda
                if col_idx - 1 in currency_indices:
                    cell.number_format = '"$"#,##0.00'
                
                # Color de fondo para filas pares
                if row_idx % 2 == 0:
                    cell.fill = even_fill
            
            row_idx += 1
        
        # Agregar fila de totales
        row_idx += 1
        for col_idx in range(1, len(columnas) + 1):
            if col_idx - 1 in numeric_indices:
                cell = worksheet.cell(row=row_idx, column=col_idx, value=totals[col_idx - 1])
                cell.font = total_font
                cell.fill = total_fill
                
                # Aplicar formato de número
                cell.number_format = '#,##0.00'
                
                # Aplicar formato de moneda si corresponde
                if col_idx - 1 in currency_indices:
                    cell.number_format = '"$"#,##0.00'
            elif col_idx == 1:
                cell = worksheet.cell(row=row_idx, column=col_idx, value="TOTALES")
                cell.font = total_font
                cell.fill = total_fill
            else:
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = total_fill

        # Aplicar autofilter a todo el rango de datos
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}{row_idx-1}"
        
        # Congelar la primera fila para que sea visible al desplazarse
        worksheet.freeze_panes = "A2"
        
    except ValueError:
        return HttpResponse("Número de pedido inválido", status=400)

    # Guardar el libro de trabajo en el buffer
    workbook.save(output)
    output.seek(0)

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Detalles_Pedidos.xlsx"'

    return response
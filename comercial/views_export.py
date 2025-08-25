import io
import time
from collections import defaultdict
from datetime import datetime
from decimal import Decimal

from django.contrib.auth.decorators import user_passes_test, login_required
from django.db.models import Q
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook import Workbook

from .models import Pedido


def es_miembro_del_grupo(nombre_grupo):
    def es_miembro(user):
        return user.groups.filter(name=nombre_grupo).exists()

    return es_miembro

# ---------------------------------- Funcion que exporta las utilidades a Excel ---------------------------------------

@login_required
@user_passes_test(lambda u: any(es_miembro_del_grupo(grupo)(u) for grupo in ['Heavens', 'Etnico', 'Fieldex', 'Juan_Matas', 'CI_Dorado']), login_url='home')
def exportar_utilidades_excel(request):
    # Determinar a qué grupo pertenece el usuario
    grupo = None
    if es_miembro_del_grupo('Heavens')(request.user):
        grupo = 'Heavens'
    elif es_miembro_del_grupo('Etnico')(request.user):
        grupo = 'Etnico'
    elif es_miembro_del_grupo('Fieldex')(request.user):
        grupo = 'Fieldex'
    elif es_miembro_del_grupo('Juan_Matas')(request.user):
        grupo = 'Juan_Matas'
    elif es_miembro_del_grupo('CI_Dorado')(request.user):
        grupo = 'CI_Dorado'

    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()

    # Definir estilos mejorados
    header_font = Font(bold=True, color="FFFFFF", name='Arial', size=11)
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    total_font = Font(bold=True, color="FFFFFF", name='Arial', size=11)
    total_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    total_align = Alignment(horizontal="center", vertical="center")
    
    subtotal_font = Font(bold=True, name='Arial', size=10)
    subtotal_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    category_font = Font(bold=True, color="000000", name='Arial', size=10)
    category_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    date_format = 'DD-MM-YYYY'
    number_format = '$#,##0.00'
    percentage_format = '0.00%'

    # Definir el color de relleno rojo suave
    fill_red_soft = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Hoja 1: Utilidades Totales General
    worksheet1 = workbook.active
    worksheet1.title = 'Utilidades Totales General'

    # Encabezados
    columns = ['No. Pedido', 'Fecha Entrega Pedido', 'Cliente', 'Exportador', 'AWB', 'Fecha Pago Cliente', 'No Factura',
               'Valor Total Factura USD', 'Valor Pagado Cliente', 'Estado Factura', 'T Cajas Enviadas',
               'Trm Monetizacion',
               'TRM Banrep', 'Valor Utilidad USD', 'Valor Recuperación USD', 'Valor Utilidad Pesos', 'Documento Cobro Utilidad',
               'Fecha Pago Utilidad', 'Diferencia O Abono', 'Estado Utilidad', 'Cobrar Utilidad']
    
    # Ajuste automático del ancho de las columnas según su contenido
    column_widths = [10, 15, 20, 15, 12, 15, 15, 15, 15, 15, 12, 12, 12, 15, 15, 15, 20, 15, 15, 15, 10]
    
    # Aplicar encabezados y ancho de columnas
    for col_num, (column_title, width) in enumerate(zip(columns, column_widths), start=1):
        cell = worksheet1.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        worksheet1.column_dimensions[worksheet1.cell(row=1, column=col_num).column_letter].width = width

    # Crear diccionarios para almacenar los totales por estado y exportadora
    totales_por_utilidad_usd = defaultdict(Decimal)
    totales_por_recuperacion_usd = defaultdict(Decimal)
    totales_por_estado = {
        'Pendiente Pago Cliente': defaultdict(Decimal),
        'Factura en abono': defaultdict(Decimal),
        'Por Facturar': defaultdict(Decimal),
        'Facturada': defaultdict(Decimal),
        'Pagada': defaultdict(Decimal)
    }
    
    # Totales generales para el resumen
    total_general_utilidad_usd = Decimal('0.00')
    total_general_recuperacion_usd = Decimal('0.00')
    total_general_por_estado = {
        'Pendiente Pago Cliente': Decimal('0.00'),
        'Factura en abono': Decimal('0.00'),
        'Por Facturar': Decimal('0.00'),
        'Facturada': Decimal('0.00'),
        'Pagada': Decimal('0.00')
    }
    total_cajas_enviadas = 0
    clientes_unicos = set()
    exportadoras_unicas = set()

    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Construir el filtro base según las fechas
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtro para las fechas
        queryset_filter = Q(fecha_entrega__gte=fecha_inicial, fecha_entrega__lte=fecha_final)
    else:
        queryset_filter = Q()  # Filtro vacío (selecciona todos)

    # Aplicar el filtro por grupo/exportadora
    if grupo == 'Heavens':
        # Si es Heavens, muestra TODOS los registros (de todas las exportadoras)
        queryset = Pedido.objects.filter(queryset_filter)
    elif grupo:
        # Para otros grupos, filtrar solo los pedidos de su exportadora
        queryset = Pedido.objects.filter(queryset_filter & Q(exportadora__nombre=grupo))
    else:
        # Si no tiene grupo específico (no debería ocurrir por el decorador)
        queryset = Pedido.objects.filter(queryset_filter)

    # Obtener los datos y acumular estadísticas
    row_num = 2  # Comenzar después del encabezado
    for pedido in queryset:
        valor_utilidad_usd = pedido.valor_total_utilidad_usd or Decimal('0.00')
        valor_recuperacion_usd = pedido.valor_total_recuperacion_usd or Decimal('0.00')
        
        # Acumular estadísticas
        exportadora = pedido.exportadora.nombre
        exportadoras_unicas.add(exportadora)
        clientes_unicos.add(pedido.cliente.nombre)
        total_cajas_enviadas += pedido.total_cajas_enviadas or 0
        
        # Calcular totales por exportadora
        totales_por_utilidad_usd[exportadora] += valor_utilidad_usd
        totales_por_recuperacion_usd[exportadora] += valor_recuperacion_usd
        total_general_utilidad_usd += valor_utilidad_usd
        total_general_recuperacion_usd += valor_recuperacion_usd
        
        # Clasificar utilidades según su estado
        estado = pedido.estado_utilidad
        if estado in totales_por_estado:
            totales_por_estado[estado][exportadora] += valor_utilidad_usd
            total_general_por_estado[estado] += valor_utilidad_usd

        # Determinar si se debe cobrar la utilidad
        cobrar_utilidad = "Sí" if pedido.estado_utilidad == "Por Facturar" or pedido.estado_utilidad == "Facturada" else "No"
        
        # Agregar fila de datos
        row = [
            pedido.pk,
            pedido.fecha_entrega,
            pedido.cliente.nombre,
            exportadora,
            pedido.awb,
            pedido.fecha_pago,
            pedido.numero_factura,
            pedido.valor_total_factura_usd,
            pedido.valor_pagado_cliente_usd,
            pedido.estado_factura,
            pedido.total_cajas_enviadas,
            pedido.trm_monetizacion,
            pedido.tasa_representativa_usd_diaria,
            valor_utilidad_usd,
            valor_recuperacion_usd,
            pedido.valor_utilidad_pesos,
            pedido.documento_cobro_utilidad,
            pedido.fecha_pago_utilidad,
            pedido.diferencia_por_abono,
            pedido.estado_utilidad,
            cobrar_utilidad,
        ]
        
        for col_num, cell_value in enumerate(row, start=1):
            cell = worksheet1.cell(row=row_num, column=col_num, value=cell_value)
            cell.border = Side(style='thin', color="DDDDDD")
            
            # Aplicar formato a fechas
            if col_num in [2, 6, 18]:  # Columnas de fechas
                if cell_value:
                    cell.number_format = date_format
            
            # Aplicar formato de moneda a las columnas específicas
            if col_num in [8, 9, 12, 13, 14, 15, 16, 19]:
                if cell_value:
                    cell.number_format = number_format
                    
            # Pintar la fila si el numero_factura es 'Pedido Cancelado'
            if pedido.numero_factura == 'Pedido Cancelado':
                cell.fill = fill_red_soft
        
        row_num += 1

    # Añadir totales al final de la primera hoja
    row_num += 2
    total_row = row_num
    worksheet1.cell(row=total_row, column=1, value="TOTALES GENERALES")
    worksheet1.cell(row=total_row, column=1).font = total_font
    worksheet1.cell(row=total_row, column=1).fill = total_fill
    worksheet1.cell(row=total_row, column=1).border = thin_border
    worksheet1.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=13)
    
    # Total de Utilidades USD
    cell = worksheet1.cell(row=total_row, column=14, value=total_general_utilidad_usd)
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    cell.number_format = number_format
    
    # Total de Recuperación USD
    cell = worksheet1.cell(row=total_row, column=15, value=total_general_recuperacion_usd)
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    cell.number_format = number_format
    
    # Total de Utilidades en Pesos
    total_pesos = sum(p.valor_utilidad_pesos or Decimal('0.00') for p in queryset)
    cell = worksheet1.cell(row=total_row, column=16, value=total_pesos)
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    cell.number_format = number_format

    # Hoja 2: Totales por Exportadora con formato mejorado
    worksheet2 = workbook.create_sheet(title='Resumen por Exportadora')
    
    # Título del informe - Ajustado según el grupo
    title_text = "RESUMEN DE UTILIDADES"
    if grupo and grupo != 'Heavens':
        title_text += f" PARA {grupo.upper()}"
    else:
        title_text += " POR EXPORTADORA"
        
    title_cell = worksheet2.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=14, name='Arial')
    worksheet2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)  # Updated to include new column
    title_cell.alignment = Alignment(horizontal="center")
    
    # Período del informe
    if fecha_inicial_str and fecha_final_str:
        period_text = f"Período: {fecha_inicial_str} al {fecha_final_str}"
        period_cell = worksheet2.cell(row=2, column=1, value=period_text)
        period_cell.font = Font(italic=True, size=10)
        worksheet2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)  # Updated to include new column
        period_cell.alignment = Alignment(horizontal="center")
        start_row = 4
    else:
        start_row = 3
    
    # Encabezados para la segunda hoja - Separando "Por Facturar" y "Facturada"
    headers = ["Exportadora", "Total Utilidades USD", "Total Recuperación USD", 
               "Pendiente Pago Cliente", "Factura en Abono", "Por Facturar", "Facturada", "Pagada"]
    header_widths = [30, 20, 20, 20, 20, 20, 20, 20]
    
    for col_num, (header, width) in enumerate(zip(headers, header_widths), start=1):
        cell = worksheet2.cell(row=start_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        worksheet2.column_dimensions[worksheet2.cell(row=start_row, column=col_num).column_letter].width = width
    
    # Agregar totales a la segunda hoja
    row_num = start_row + 1
    for exportadora in sorted(totales_por_utilidad_usd.keys()):
        worksheet2.cell(row=row_num, column=1, value=exportadora).border = thin_border
        
        cell = worksheet2.cell(row=row_num, column=2, value=totales_por_utilidad_usd[exportadora])
        cell.number_format = number_format
        cell.border = thin_border
        
        cell = worksheet2.cell(row=row_num, column=3, value=totales_por_recuperacion_usd[exportadora])
        cell.number_format = number_format
        cell.border = thin_border
        
        # Agregar valores por estado
        cell = worksheet2.cell(row=row_num, column=4, value=totales_por_estado['Pendiente Pago Cliente'][exportadora])
        cell.number_format = number_format
        cell.border = thin_border
        
        cell = worksheet2.cell(row=row_num, column=5, value=totales_por_estado['Factura en abono'][exportadora])
        cell.number_format = number_format
        cell.border = thin_border
        
        # Separar "Por Facturar" y "Facturada"
        cell = worksheet2.cell(row=row_num, column=6, value=totales_por_estado['Por Facturar'][exportadora])
        cell.number_format = number_format
        cell.border = thin_border
        
        cell = worksheet2.cell(row=row_num, column=7, value=totales_por_estado['Facturada'][exportadora])
        cell.number_format = number_format
        cell.border = thin_border
        
        cell = worksheet2.cell(row=row_num, column=8, value=totales_por_estado['Pagada'][exportadora])
        cell.number_format = number_format
        cell.border = thin_border
        
        row_num += 1

    # Agregar totales generales
    row_num += 1
    totals_row = row_num
    
    cell = worksheet2.cell(row=totals_row, column=1, value="TOTAL GENERAL")
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    
    cell = worksheet2.cell(row=totals_row, column=2, value=total_general_utilidad_usd)
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    cell.number_format = number_format
    
    cell = worksheet2.cell(row=totals_row, column=3, value=total_general_recuperacion_usd)
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    cell.number_format = number_format
    
    cell = worksheet2.cell(row=totals_row, column=4, value=total_general_por_estado['Pendiente Pago Cliente'])
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    cell.number_format = number_format
    
    cell = worksheet2.cell(row=totals_row, column=5, value=total_general_por_estado['Factura en abono'])
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    cell.number_format = number_format
    
    # Separar "Por Facturar" y "Facturada" en totales generales
    cell = worksheet2.cell(row=totals_row, column=6, value=total_general_por_estado['Por Facturar'])
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    cell.number_format = number_format
    
    cell = worksheet2.cell(row=totals_row, column=7, value=total_general_por_estado['Facturada'])
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    cell.number_format = number_format
    
    cell = worksheet2.cell(row=totals_row, column=8, value=total_general_por_estado['Pagada'])
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    cell.number_format = number_format

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                           content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    # Nombre del archivo con fecha si se especificó un rango
    filename_parts = []
    if grupo and grupo != 'Heavens':
        filename_parts.append(f"{grupo}")
    filename_parts.append("utilidades_pedidos")
    
    if fecha_inicial_str and fecha_final_str:
        filename_parts.append(f"{fecha_inicial_str}_a_{fecha_final_str}")
    
    filename = "_".join(filename_parts) + ".xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


# ---------------------------------- Funcion que exporta las utilidades a Excel ---------------------------------------

# ---------------------------------- Funcion que exporta los pedidos a Excel ------------------------------------------

@login_required
@user_passes_test(lambda u: any(es_miembro_del_grupo(grupo)(u) for grupo in ['Heavens', 'Etnico', 'Fieldex', 'Juan_Matas', 'CI_Dorado']), login_url='home')
def exportar_pedidos_excel_general(request):
    from django.core.paginator import Paginator
    from django.db import connection
    import gc
    
    # Inicializar control de tiempo para timeout
    start_time = time.time()
    max_execution_time = 300  # 5 minutos máximo
    
    # Determinar a qué grupo pertenece el usuario
    grupo = None
    if es_miembro_del_grupo('Heavens')(request.user):
        grupo = 'Heavens'
    elif es_miembro_del_grupo('Etnico')(request.user):
        grupo = 'Etnico'
    elif es_miembro_del_grupo('Fieldex')(request.user):
        grupo = 'Fieldex'
    elif es_miembro_del_grupo('Juan_Matas')(request.user):
        grupo = 'Juan_Matas'
    elif es_miembro_del_grupo('CI_Dorado')(request.user):
        grupo = 'CI_Dorado'
        
    # 1. Crear un buffer en memoria para almacenar el archivo Excel
    output = io.BytesIO()

    # 2. Crear el Workbook en modo write_only para optimizar el uso de memoria
    workbook = Workbook(write_only=True)
    ws = workbook.create_sheet(title='Pedidos')

    # 3. Definir estilos para los encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e0c42", end_color="1e0c42", fill_type="solid")
    sub_header_fill = PatternFill(start_color="0B6FA4", end_color="0B6FA4", fill_type="solid")

    # 4. Definir los encabezados para Pedido y DetallePedido
    pedido_headers = [
        'No', 'Cliente', 'Semana', 'Fecha Solicitud', 'Fecha Entrega', 'Fecha Llegada',
        'Exportador', 'Subexportadora', 'Intermediario', 'Dias Cartera', 'AWB',
        'Destino', 'Aerolinea', 'Agencia De Carga', 'Responsable Reserva', 'Numero Factura',
        'Total Cajas Solicitadas', 'Total Cajas Enviadas', 'Peso Bruto Solicitado',
        'Peso Bruto Enviado', 'Pallets Solicitados', 'Pallets Enviados', 'Peso AWB',
        'ETA', 'ETD', 'Variedades', 'Descuento Comercial', 'No NC', 'Motivo NC',
        'Valor Total NC', 'Valor Pagado Cliente', 'Estado Factura', 'Utilidad Bancaria USD',
        'Fecha Pago Cliente', 'TRM Monetización', 'Fecha Monetización', 'Trm Banrep',
        'Trm Cotización', 'Diferencia Pago', 'Dias Vencimiento', 'Valor Total Factura USD',
        'Valor Utilidad USD', 'Valor Utilidad Pesos', 'Valor Recuperacion USD', 'Documento Cobro Utilidad',
        'Fecha Pago Utilidad', 'Estado Utilidad', 'Estado Cancelacion', 'Estado Documentos',
        'Estado Reserva', 'Termo', 'Diferencia AWB/Factura', 'Eta Real', 'Estado Pedido',
        'Observaciones Tracking', 'Observaciones Generales'
    ]

    detalle_headers = [
        'Pedido', 'F Entrega', 'Exportador', 'Cliente', 'Fruta', 'Presentacion', 'Cajas Solicitadas',
        'Peso Presentacion', 'kilos', 'Cajas Enviadas', 'Kilos Enviados', 'Diferencia', 'Tipo Caja',
        'Referencia', 'Stiker', 'Lleva Contenedor', 'Ref Contenedor', 'Cant Contenedor', 'Tarifa utilidad', 'Tarifa Recuperacion',
        'Valor x Caja USD', 'Valor X Producto', 'No Cajas NC', 'Valor NC', 'Afecta utilidad',
        'Valor Total utilidad Producto', 'Valor Total Recuperacion X Producto', 'Precio Proforma', 'Observaciones'
    ]

    # 5. Verificar si el usuario incluyó detalles
    incluir_detalles = request.POST.get('incluir_detalles') == 'true'

    # 6. Obtener filtros de fecha desde el formulario (si existen)
    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # 7. Filtrar los pedidos según fechas con optimizaciones
    try:
        if (fecha_inicial_str and fecha_final_str):
            fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
            fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')
            base_filter = Q(fecha_entrega__gte=fecha_inicial, fecha_entrega__lte=fecha_final)
        else:
            base_filter = Q()
            
        # Filtrar según el grupo del usuario
        if grupo == 'Heavens':
            # Si es Heavens, muestra TODOS los registros (de todas las exportadoras)
            pedidos_qs = Pedido.objects.filter(base_filter)
        elif grupo:
            # Para otros grupos, filtrar solo los pedidos de su exportadora
            pedidos_qs = Pedido.objects.filter(base_filter & Q(exportadora__nombre=grupo))
        else:
            # Si no tiene grupo específico (no debería ocurrir por el decorador)
            pedidos_qs = Pedido.objects.filter(base_filter)
            
        # Optimización crítica: usar select_related para evitar N+1 queries
        pedidos_qs = pedidos_qs.select_related(
            'cliente',
            'exportadora', 
            'subexportadora',
            'intermediario',
            'destino',
            'aerolinea',
            'agencia_carga',
            'responsable_reserva'
        ).order_by('id')  # Orden consistente para paginación
        
        # Contar total de registros para validar límites
        total_pedidos = pedidos_qs.count()
        
        # Límite de seguridad: máximo 10,000 pedidos por exportación
        if total_pedidos > 10000:
            return HttpResponse(
                f"Demasiados registros ({total_pedidos}). Por favor, use filtros de fecha para reducir el conjunto de datos a menos de 10,000 registros.",
                status=400
            )
            
        # Usar paginación para procesar en lotes de 100 pedidos
        paginator = Paginator(pedidos_qs, 100)
            
    except ValueError:
        return HttpResponse("Fecha inválida", status=400)

    # Escribir encabezados una sola vez al inicio (optimizado)
    if incluir_detalles:
        # Encabezados combinados para pedidos y detalles
        combined_headers = pedido_headers + ['---DETALLES---'] + detalle_headers
        # Usar lista simple para encabezados, aplicar formato después
        ws.append(combined_headers)
    else:
        # Solo encabezados de pedidos
        ws.append(pedido_headers)
    
    # Procesar pedidos en lotes para optimizar memoria
    for page_num in range(1, paginator.num_pages + 1):
        page = paginator.page(page_num)
        
        # Prefetch detalles si es necesario
        if incluir_detalles:
            # Obtener IDs de pedidos de la página actual
            pedido_ids = [p.id for p in page.object_list]
            
            # Hacer una consulta optimizada con prefetch_related
            pedidos_con_detalles = Pedido.objects.filter(
                id__in=pedido_ids
            ).select_related(
                'cliente',
                'exportadora', 
                'subexportadora',
                'intermediario',
                'destino',
                'aerolinea',
                'agencia_carga',
                'responsable_reserva'
            ).prefetch_related(
                'detallepedido_set__fruta',
                'detallepedido_set__presentacion', 
                'detallepedido_set__tipo_caja',
                'detallepedido_set__referencia'
            ).order_by('id')
        else:
            pedidos_con_detalles = page.object_list
            
        for pedido in pedidos_con_detalles:

            # Escribir los datos del Pedido
            pedido_data = [
                pedido.pk,
                pedido.cliente.nombre if pedido.cliente else '',
                pedido.semana,
                pedido.fecha_solicitud.strftime('%Y-%m-%d') if pedido.fecha_solicitud else '',
                pedido.fecha_entrega.strftime('%Y-%m-%d') if pedido.fecha_entrega else '',
                pedido.fecha_llegada.strftime('%Y-%m-%d') if pedido.fecha_llegada else '',
                pedido.exportadora.nombre if pedido.exportadora else '',
                pedido.subexportadora.nombre if pedido.subexportadora else '',
                pedido.intermediario.nombre if pedido.intermediario else '',
                pedido.dias_cartera,
                pedido.awb,
                pedido.destino.codigo if pedido.destino else '',
                pedido.aerolinea.nombre if pedido.aerolinea else '',
                pedido.agencia_carga.nombre if pedido.agencia_carga else '',
                pedido.responsable_reserva.nombre if pedido.responsable_reserva else '',
                pedido.numero_factura,
                pedido.total_cajas_solicitadas,
                pedido.total_cajas_enviadas,
                pedido.total_peso_bruto_solicitado,
                pedido.total_peso_bruto_enviado,
                pedido.total_piezas_solicitadas,
                pedido.total_piezas_enviadas,
                pedido.peso_awb,
                pedido.eta.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S') if pedido.eta else '',
                pedido.etd.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S') if pedido.etd else '',
                pedido.variedades,
                pedido.descuento,
                pedido.nota_credito_no,
                pedido.motivo_nota_credito,
                pedido.valor_total_nota_credito_usd,
                pedido.valor_pagado_cliente_usd,
                pedido.estado_factura,
                pedido.utilidad_bancaria_usd,
                pedido.fecha_pago.strftime('%Y-%m-%d') if pedido.fecha_pago else '',
                pedido.trm_monetizacion,
                pedido.fecha_monetizacion.strftime('%Y-%m-%d') if pedido.fecha_monetizacion else '',
                pedido.tasa_representativa_usd_diaria,
                pedido.trm_cotizacion,
                pedido.diferencia_por_abono,
                pedido.dias_de_vencimiento,
                pedido.valor_total_factura_usd,
                pedido.valor_total_utilidad_usd,
                pedido.valor_utilidad_pesos,
                pedido.valor_total_recuperacion_usd,
                pedido.documento_cobro_utilidad,
                pedido.fecha_pago_utilidad.strftime('%Y-%m-%d') if pedido.fecha_pago_utilidad else '',
                pedido.estado_utilidad,
                pedido.estado_cancelacion,
                pedido.estado_documentos,
                pedido.estatus_reserva,
                pedido.termo,
                pedido.diferencia_peso_factura_awb,
                pedido.eta_real.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S') if pedido.eta_real else '',
                pedido.estado_pedido,
                pedido.observaciones_tracking,
                pedido.observaciones
            ]
            
            if not incluir_detalles:
                # Si no incluimos detalles, escribir solo la fila del pedido
                ws.append(pedido_data)
            else:
                # Si incluimos detalles, obtener los detalles del pedido (ya prefetched)
                detalles = pedido.detallepedido_set.all()
                
                if detalles:
                    # Escribir una fila por cada detalle
                    for detalle in detalles:
                        detalle_data = [
                            detalle.pedido.pk,
                            detalle.pedido.fecha_entrega.strftime('%Y-%m-%d') if detalle.pedido.fecha_entrega else '',
                            detalle.pedido.exportadora.nombre if detalle.pedido.exportadora else '',
                            detalle.pedido.cliente.nombre if detalle.pedido.cliente else '',
                            detalle.fruta.nombre if detalle.fruta else '',
                            detalle.presentacion.nombre if detalle.presentacion else '',
                            detalle.cajas_solicitadas,
                            detalle.presentacion_peso,
                            detalle.kilos,
                            detalle.cajas_enviadas,
                            detalle.kilos_enviados,
                            detalle.diferencia,
                            detalle.tipo_caja.nombre if detalle.tipo_caja else '',
                            detalle.referencia.nombre if detalle.referencia else '',
                            detalle.stickers,
                            detalle.lleva_contenedor,
                            detalle.referencia_contenedor,
                            detalle.cantidad_contenedores,
                            detalle.tarifa_utilidad,
                            detalle.tarifa_recuperacion,
                            detalle.valor_x_caja_usd,
                            detalle.valor_x_producto,
                            detalle.valor_total_recuperacion_x_producto,
                            detalle.no_cajas_nc,
                            detalle.valor_nota_credito_usd,
                            detalle.afecta_utilidad,
                            detalle.valor_total_utilidad_x_producto,
                            detalle.precio_proforma,
                            detalle.observaciones,
                        ]
                        # Combinar datos del pedido con datos del detalle
                        combined_row = pedido_data + ['---DETALLES---'] + detalle_data
                        ws.append(combined_row)
                else:
                    # Si no hay detalles, escribir solo la fila del pedido con columnas vacías para detalles
                    empty_detalle = [''] * len(detalle_headers)
                    combined_row = pedido_data + ['---DETALLES---'] + empty_detalle
                    ws.append(combined_row)
        
        # Verificar timeout cada lote
        current_time = time.time()
        if current_time - start_time > max_execution_time:
            return HttpResponse(
                f"Tiempo de ejecución excedido ({max_execution_time} segundos). Por favor, use filtros de fecha más específicos para reducir el conjunto de datos.",
                status=408  # Request Timeout
            )
        
        # Liberar memoria después de cada lote
        if page_num % 5 == 0:  # Cada 5 lotes (500 pedidos)
            gc.collect()
            
        # Verificar conexiones de base de datos para evitar timeouts
        connection.close()
    # 9. Guardar y retornar el archivo
    workbook.save(output)
    output.seek(0)
    
    # Personalizar el nombre del archivo según el grupo
    filename_parts = []
    if grupo and grupo != 'Heavens':
        filename_parts.append(f"{grupo}")
    filename_parts.append("pedidos")
    
    if fecha_inicial_str and fecha_final_str:
        filename_parts.append(f"{fecha_inicial_str}_a_{fecha_final_str}")
    
    filename = "_".join(filename_parts) + ".xlsx"
    
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response






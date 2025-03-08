import json
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .models import Pedido, Cliente, Exportador
from .resources import obtener_datos_con_totales_cliente


def es_miembro_del_grupo(nombre_grupo):
    def es_miembro(user):
        return user.groups.filter(name=nombre_grupo).exists()

    return es_miembro

# ------------------------------ Dashboard Cliente ---------------------------------------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Heavens'), login_url='home'))
def dashboard_cliente(request):
    # Obtener parámetros de filtro
    cliente_id = request.GET.get('cliente')
    fecha_inicial_str = request.GET.get('fecha_inicial')
    fecha_final_str = request.GET.get('fecha_final')
    grupo = request.GET.get('grupo')

    # Valores por defecto
    hoy = datetime.now().date()
    primer_dia_mes = datetime(hoy.year, hoy.month, 1).date()

    # Convertir a objetos fecha si existen
    fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d').date() if fecha_inicial_str else primer_dia_mes
    fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d').date() if fecha_final_str else hoy

    # Obtener listas para los filtros
    clientes = Cliente.objects.all().order_by('nombre')
    exportadoras = Exportador.objects.all().order_by('nombre')

    # Si no hay cliente seleccionado, mostrar solo el formulario de filtro
    if not cliente_id:
        context = {
            'clientes': clientes,
            'exportadoras': exportadoras,
            'fecha_inicial': fecha_inicial,
            'fecha_final': fecha_final,
        }
        return render(request, 'dashboard_cliente.html', context)

    # Obtener el cliente seleccionado
    cliente = Cliente.objects.get(id=cliente_id)

    # Obtener los datos
    pedidos, totales = obtener_datos_con_totales_cliente(
        fecha_inicial=fecha_inicial,
        fecha_final=fecha_final,
        cliente=cliente,
        grupo=grupo
    )

    # Calcular totales
    total_facturas = sum(total['total_factura'] for total in totales)
    total_pagado = sum(total['total_pagado'] for total in totales)
    total_utilidad = sum(total['total_utilidad'] for total in totales)
    total_notas_credito = sum(total['total_nc'] for total in totales)
    print(total_notas_credito)
    total_descuentos = sum(total['total_descuentos'] for total in totales)
    saldo_pendiente = total_facturas - total_pagado - total_utilidad - total_notas_credito - total_descuentos
    
    # Add JSON serialized chart data to ensure precision
    chart_data = {
        'total_pagado': float(total_pagado),
        'saldo_pendiente': float(saldo_pendiente),
        'total_notas_credito': float(total_notas_credito),
        'total_descuentos': float(total_descuentos),
        'total_utilidad': float(total_utilidad)
    }
    
    chart_data_json = json.dumps(chart_data)
    
    # Calcular total de facturas vencidas
    total_facturas_vencidas = 0
    for p in pedidos:
        if p['estado_factura'] != 'Pagada' and p['fecha_esperada_pago']:
            fecha_esperada = p['fecha_esperada_pago'].date() if isinstance(p['fecha_esperada_pago'], datetime) else p['fecha_esperada_pago']
            if fecha_esperada < hoy:
                saldo = (p['valor_total_factura_usd'] - p['valor_pagado_cliente_usd'] - 
                        p['utilidad_bancaria_usd'] - p['valor_total_nota_credito_usd'] - p['descuento'])
                total_facturas_vencidas += saldo

    # Preparar datos para los gráficos
    # 1. Tendencia de pagos (usar fechas seleccionadas, si no hay datos mostrar último año)
    pagos_recientes = Pedido.objects.filter(
        cliente=cliente,
        fecha_pago__isnull=False,
        fecha_pago__gte=fecha_inicial,
        fecha_pago__lte=fecha_final
    ).order_by('fecha_pago').values(
        'fecha_pago', 'valor_pagado_cliente_usd', 'numero_factura', 'fecha_entrega'
    )
    
    # Si no hay pagos en el rango seleccionado, mostrar último año
    if not pagos_recientes:
        un_ano_atras = hoy - timedelta(days=365)
        pagos_recientes = Pedido.objects.filter(
            cliente=cliente,
            fecha_pago__isnull=False,
            fecha_pago__gte=un_ano_atras
        ).order_by('fecha_pago').values(
            'fecha_pago', 'valor_pagado_cliente_usd', 'numero_factura', 'fecha_entrega'
        )

    pagos_fechas = []
    pagos_montos = []
    dias_de_pago = []

    for p in pagos_recientes:
        if p['fecha_entrega'] and p['fecha_pago']:
            pagos_fechas.append(p['fecha_pago'].strftime('%Y-%m-%d'))
            pagos_montos.append(float(p['valor_pagado_cliente_usd']))
            dias = (p['fecha_pago'] - p['fecha_entrega']).days
            dias_de_pago.append(dias)

    # Datos para gráficos
    pagos_fechas_json = json.dumps(pagos_fechas)
    pagos_montos_json = json.dumps(pagos_montos)
    dias_de_pago_json = json.dumps(dias_de_pago)
    dias_cartera_cliente = cliente.negociaciones_cartera
    dias_cartera_json = json.dumps([dias_cartera_cliente] * len(pagos_fechas))

    # 2. Facturas por exportadora
    exportadoras_data = {}
    for p in pedidos:
        exp_nombre = p['exportadora__nombre']
        if exp_nombre in exportadoras_data:
            exportadoras_data[exp_nombre] += float(p['valor_total_factura_usd'])
        else:
            exportadoras_data[exp_nombre] = float(p['valor_total_factura_usd'])

    exportadoras_nombres = json.dumps(list(exportadoras_data.keys()))
    exportadoras_montos = json.dumps(list(exportadoras_data.values()))

    # 3. Facturas próximas a vencer
    facturas_proximas = []
    hoy = datetime.now().date()

    for p in pedidos:
        if p['estado_factura'] != 'Pagada' and p['fecha_esperada_pago']:
            fecha_esperada = p['fecha_esperada_pago'].date() if isinstance(p['fecha_esperada_pago'], datetime) else p[
                'fecha_esperada_pago']
            dias_restantes = (fecha_esperada - hoy).days

            if dias_restantes >= -5 and dias_restantes <= 20:  # Mostrar facturas vencidas hasta 5 días y por vencer hasta 20 días
                facturas_proximas.append({
                    'numero_factura': p['numero_factura'],
                    'valor_total_factura_usd': p['valor_total_factura_usd'],
                    'fecha_entrega': p['fecha_entrega'],
                    'fecha_esperada_pago': fecha_esperada,
                    'dias_restantes': dias_restantes
                })

    # Ordenar por días restantes (las más urgentes primero)
    facturas_proximas.sort(key=lambda x: x['dias_restantes'])

    # Pagos más recientes para la línea de tiempo
    pagos_recientes_timeline = Pedido.objects.filter(
        cliente=cliente,
        fecha_pago__isnull=False
    ).order_by('-fecha_pago')[:8] # Últimos 8 pagos

    context = {
        'selected_cliente': cliente_id,
        'selected_grupo': grupo,
        'cliente': cliente,
        'clientes': clientes,
        'exportadoras': exportadoras,
        'fecha_inicial': fecha_inicial,
        'fecha_final': fecha_final,
        'pedidos': pedidos,
        'totales': totales,
        'total_facturas': total_facturas,
        'total_pagado': total_pagado,
        'total_utilidad': total_utilidad,
        'total_notas_credito': total_notas_credito,
        'total_descuentos': total_descuentos,
        'saldo_pendiente': saldo_pendiente,
        'total_facturas_vencidas': total_facturas_vencidas,
        'pagos_fechas': pagos_fechas_json,
        'pagos_montos': pagos_montos_json,
        'dias_de_pago': dias_de_pago_json,
        'dias_cartera': dias_cartera_json,
        'exportadoras_nombres': exportadoras_nombres,
        'exportadoras_montos': exportadoras_montos,
        'facturas_proximas': facturas_proximas,
        'pagos_recientes': pagos_recientes_timeline,
        'chart_data_json': chart_data_json,
    }

    return render(request, 'dashboard_cliente.html', context)


# ------------------------------ Dashboard Cliente ---------------------------------------------------------------

# ------------------------------ Exportar Cartera O Estado De Cuenta Dashboard Cliente----------------------------

@login_required
@user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home'))
def exportar_cartera_cliente_dashboard(request):
    cliente_id = request.GET.get('cliente', '')
    fecha_inicial_str = request.GET.get('fecha_inicial', '')
    fecha_final_str = request.GET.get('fecha_final', '')
    grupo = request.GET.get('grupo', '')

    hoy = datetime.now().date()
    primer_dia_mes = datetime(hoy.year, hoy.month, 1).date()

    fecha_inicial = parse_date(fecha_inicial_str) or primer_dia_mes
    fecha_final = parse_date(fecha_final_str) or hoy

    if not cliente_id:
        return redirect('dashboard_cliente')

    # Obtener el cliente seleccionado
    cliente = Cliente.objects.get(id=cliente_id)

    # Obtener los datos
    pedidos, totales = obtener_datos_con_totales_cliente(
        fecha_inicial=fecha_inicial,
        fecha_final=fecha_final,
        cliente=cliente,
        grupo=grupo
    )

    # Calcular totales generales
    total_facturas = sum(total['total_factura'] for total in totales)
    total_pagado = sum(total['total_pagado'] for total in totales)
    total_utilidad = sum(total['total_utilidad'] for total in totales)
    total_notas_credito = sum(total['total_nc'] for total in totales)
    total_descuentos = sum(total['total_descuentos'] for total in totales)
    saldo_pendiente = total_facturas - total_pagado - total_utilidad - total_notas_credito - total_descuentos
    
    # Calcular total de facturas vencidas
    total_facturas_vencidas = 0
    for p in pedidos:
        if p['estado_factura'] != 'Pagada' and p['fecha_esperada_pago']:
            fecha_esperada = p['fecha_esperada_pago'].date() if isinstance(p['fecha_esperada_pago'], datetime) else p['fecha_esperada_pago']
            if fecha_esperada < hoy:
                saldo = (p['valor_total_factura_usd'] - p['valor_pagado_cliente_usd'] - 
                        p['utilidad_bancaria_usd'] - p['valor_total_nota_credito_usd'] - p['descuento'])
                total_facturas_vencidas += saldo

    # Crear workbook de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Cartera_Cliente_{cliente.nombre.replace(" ", "_")}_{fecha_inicial.strftime("%Y%m%d")}_{fecha_final.strftime("%Y%m%d")}.xlsx'
    
    workbook = Workbook()
    
    # Hoja de resumen
    ws_resumen = workbook.active
    ws_resumen.title = "Resumen"
    
    # Estilos
    title_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=11)
    
    title_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    total_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
    warning_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    currency_format = '$#,##0.00'
    
    # Información del cliente
    ws_resumen.merge_cells('A1:G1')
    ws_resumen['A1'] = f'ESTADO DE CUENTA - {cliente.nombre.upper()}'
    ws_resumen['A1'].font = title_font
    ws_resumen['A1'].fill = title_fill
    ws_resumen['A1'].alignment = align_center
    
    ws_resumen['A3'] = 'Cliente:'
    ws_resumen['B3'] = cliente.nombre
    ws_resumen['A4'] = 'Dirección:'
    ws_resumen['B4'] = cliente.direccion
    ws_resumen['A5'] = 'Tax ID:'
    ws_resumen['B5'] = cliente.tax_id
    ws_resumen['A6'] = 'Días de cartera:'
    ws_resumen['B6'] = cliente.negociaciones_cartera
    ws_resumen['A7'] = 'País:'
    ws_resumen['B7'] = cliente.destino_iata.pais if cliente.destino_iata else ''
    
    ws_resumen['D3'] = 'Fecha inicial:'
    ws_resumen['E3'] = fecha_inicial
    ws_resumen['D4'] = 'Fecha final:'
    ws_resumen['E4'] = fecha_final
    
    # Aplicar estilo a las celdas de información
    for row in range(3, 8):
        for col in ['A', 'D']:
            ws_resumen[f'{col}{row}'].font = Font(name='Calibri', size=11, bold=True)
    
    # Resumen financiero
    ws_resumen['A9'] = 'RESUMEN FINANCIERO'
    ws_resumen.merge_cells('A9:G9')
    ws_resumen['A9'].font = title_font
    ws_resumen['A9'].fill = title_fill
    ws_resumen['A9'].alignment = align_center
    
    ws_resumen['A10'] = 'Total Facturado (USD)'
    ws_resumen['B10'] = total_facturas
    ws_resumen['B10'].number_format = currency_format
    
    ws_resumen['A11'] = 'Total Pagado (USD)'
    ws_resumen['B11'] = total_pagado
    ws_resumen['B11'].number_format = currency_format
    
    ws_resumen['A12'] = 'Total Utilidad Bancaria (USD)'
    ws_resumen['B12'] = total_utilidad
    ws_resumen['B12'].number_format = currency_format
    
    ws_resumen['A13'] = 'Total Notas de Crédito (USD)'
    ws_resumen['B13'] = total_notas_credito
    ws_resumen['B13'].number_format = currency_format
    
    ws_resumen['A14'] = 'Total Descuentos (USD)'
    ws_resumen['B14'] = total_descuentos
    ws_resumen['B14'].number_format = currency_format
    
    ws_resumen['A15'] = 'Total Facturas Vencidas (USD)'
    ws_resumen['B15'] = total_facturas_vencidas
    ws_resumen['B15'].number_format = currency_format
    ws_resumen['A15'].fill = warning_fill
    ws_resumen['B15'].fill = warning_fill
    ws_resumen['A15'].font = Font(name='Calibri', size=11, bold=True, color='9B2C2C')
    ws_resumen['B15'].font = Font(name='Calibri', size=11, bold=True, color='9B2C2C')
    
    ws_resumen['A16'] = 'Saldo Pendiente (USD)'
    ws_resumen['B16'] = saldo_pendiente
    ws_resumen['B16'].font = Font(name='Calibri', size=12, bold=True)
    ws_resumen['B16'].number_format = currency_format
    ws_resumen['A16'].fill = total_fill
    ws_resumen['B16'].fill = total_fill
    
    # Aplicar estilo a las celdas del resumen financiero
    for row in range(10, 15):
        ws_resumen[f'A{row}'].font = data_font
    
    # Resumen por exportadora
    ws_resumen['A18'] = 'RESUMEN POR EXPORTADORA'
    ws_resumen.merge_cells('A18:I18')
    ws_resumen['A18'].font = title_font
    ws_resumen['A18'].fill = title_fill
    ws_resumen['A18'].alignment = align_center
    
    # Encabezados
    headers = ['Intermediario', 'Cliente', 'Exportadora', 'Total Facturas', 'Total Pagado', 'Utilidad Bancaria', 'Notas Crédito', 'Descuentos', 'Saldo']
    for col_idx, header in enumerate(headers):
        col_letter = get_column_letter(col_idx + 1)
        ws_resumen[f'{col_letter}19'] = header
        ws_resumen[f'{col_letter}19'].font = header_font
        ws_resumen[f'{col_letter}19'].fill = header_fill
        ws_resumen[f'{col_letter}19'].alignment = align_center
        ws_resumen[f'{col_letter}19'].border = border
    
    # Datos por exportadora
    row_idx = 20
    for total in totales:
        saldo = total['total_factura'] - total['total_pagado'] - total['total_utilidad'] - total['total_nc'] - total['total_descuentos']
        
        ws_resumen[f'A{row_idx}'] = total['intermediario__nombre'] if total['intermediario__nombre'] else '-'
        ws_resumen[f'B{row_idx}'] = total['cliente__nombre']
        ws_resumen[f'C{row_idx}'] = total['exportadora__nombre']
        ws_resumen[f'D{row_idx}'] = total['total_factura']
        ws_resumen[f'D{row_idx}'].number_format = currency_format
        ws_resumen[f'E{row_idx}'] = total['total_pagado']
        ws_resumen[f'E{row_idx}'].number_format = currency_format
        ws_resumen[f'F{row_idx}'] = total['total_utilidad']
        ws_resumen[f'F{row_idx}'].number_format = currency_format
        ws_resumen[f'G{row_idx}'] = total['total_nc']
        ws_resumen[f'G{row_idx}'].number_format = currency_format
        ws_resumen[f'H{row_idx}'] = total['total_descuentos']
        ws_resumen[f'H{row_idx}'].number_format = currency_format
        ws_resumen[f'I{row_idx}'] = saldo
        ws_resumen[f'I{row_idx}'].number_format = currency_format
        
        # Aplicar formato a la fila
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws_resumen[f'{col_letter}{row_idx}'].font = data_font
            ws_resumen[f'{col_letter}{row_idx}'].border = border
            if col_idx >= 4:  # Para columnas numéricas
                ws_resumen[f'{col_letter}{row_idx}'].alignment = align_right
            else:
                ws_resumen[f'{col_letter}{row_idx}'].alignment = align_left
        
        row_idx += 1
    
    # Ajustar anchos de columna
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws_resumen.column_dimensions[col_letter].width = 18
    
    # Segunda hoja: Detalle de facturas
    ws_facturas = workbook.create_sheet(title="Facturas")
    
    # Título
    ws_facturas.merge_cells('A1:L1')
    ws_facturas['A1'] = f'DETALLE DE FACTURAS - {cliente.nombre.upper()}'
    ws_facturas['A1'].font = title_font
    ws_facturas['A1'].fill = title_fill
    ws_facturas['A1'].alignment = align_center
    
    # Encabezados de facturas
    headers_facturas = ['Factura', 'AWB', 'Fecha Entrega', 'Fecha Esperada Pago', 'Exportadora', 'Total Factura', 
                        'Total Pagado', 'Nota Crédito', 'Descuento', 'Utilidad', 'Saldo', 'Estado']
    for col_idx, header in enumerate(headers_facturas):
        col_letter = get_column_letter(col_idx + 1)
        ws_facturas[f'{col_letter}3'] = header
        ws_facturas[f'{col_letter}3'].font = header_font
        ws_facturas[f'{col_letter}3'].fill = header_fill
        ws_facturas[f'{col_letter}3'].alignment = align_center
        ws_facturas[f'{col_letter}3'].border = border
    
    # Datos de facturas
    row_idx = 4
    for pedido in pedidos:
        saldo = (pedido['valor_total_factura_usd'] - 
                pedido['valor_pagado_cliente_usd'] - 
                pedido['utilidad_bancaria_usd'] - 
                pedido['valor_total_nota_credito_usd'] - 
                pedido['descuento'])
        
        ws_facturas[f'A{row_idx}'] = pedido['numero_factura']
        ws_facturas[f'B{row_idx}'] = pedido['awb']
        ws_facturas[f'C{row_idx}'] = pedido['fecha_entrega'].strftime('%d/%m/%Y') if pedido['fecha_entrega'] else ''
        ws_facturas[f'D{row_idx}'] = pedido['fecha_esperada_pago'].strftime('%d/%m/%Y') if pedido['fecha_esperada_pago'] else ''
        ws_facturas[f'E{row_idx}'] = pedido['exportadora__nombre']
        ws_facturas[f'F{row_idx}'] = pedido['valor_total_factura_usd']
        ws_facturas[f'F{row_idx}'].number_format = currency_format
        ws_facturas[f'G{row_idx}'] = pedido['valor_pagado_cliente_usd']
        ws_facturas[f'G{row_idx}'].number_format = currency_format
        ws_facturas[f'H{row_idx}'] = pedido['valor_total_nota_credito_usd']
        ws_facturas[f'H{row_idx}'].number_format = currency_format
        ws_facturas[f'I{row_idx}'] = pedido['descuento']
        ws_facturas[f'I{row_idx}'].number_format = currency_format
        ws_facturas[f'J{row_idx}'] = pedido['utilidad_bancaria_usd']
        ws_facturas[f'J{row_idx}'].number_format = currency_format
        ws_facturas[f'K{row_idx}'] = saldo
        ws_facturas[f'K{row_idx}'].number_format = currency_format
        
        # Estado de la factura
        estado = pedido['estado_factura']
        if estado == 'Pagada':
            estado_texto = 'Pagada'
        elif estado == 'Factura sin valor':
            estado_texto = 'Sin valor'
        elif estado == 'Cancelada':
            estado_texto = 'Cancelada'
        elif pedido.get('estado_cancelacion') == 'Pendiente':  # Changed from ['estado_cancelacion'] to .get()
            estado_texto = 'Cancelación Pendiente'
        elif pedido.get('estado_cancelacion') == 'Autorizado':  # Changed from ['estado_cancelacion'] to .get()
            estado_texto = 'Cancelada'
        elif pedido.get('dias_de_vencimiento', 0) > 0:
            estado_texto = f'Vencida {pedido.get("dias_de_vencimiento", 0)} días'
        else:
            estado_texto = 'Pendiente'
        
        ws_facturas[f'L{row_idx}'] = estado_texto
        
        # Aplicar formato a la fila
        for col_idx in range(1, len(headers_facturas) + 1):
            col_letter = get_column_letter(col_idx)
            ws_facturas[f'{col_letter}{row_idx}'].font = data_font
            ws_facturas[f'{col_letter}{row_idx}'].border = border
            if col_idx in [6, 7, 8, 9, 10, 11]:  # Para columnas numéricas
                ws_facturas[f'{col_letter}{row_idx}'].alignment = align_right
            else:
                ws_facturas[f'{col_letter}{row_idx}'].alignment = align_left
        
        row_idx += 1
    
    # Ajustar anchos de columna
    for col_idx in range(1, len(headers_facturas) + 1):
        col_letter = get_column_letter(col_idx)
        ws_facturas.column_dimensions[col_letter].width = 16
    
    # Congelar paneles para mejor visualización
    ws_facturas.freeze_panes = 'A4'
    ws_resumen.freeze_panes = 'A20'
    
    # Guardar el libro de trabajo en la respuesta
    workbook.save(response)
    return response
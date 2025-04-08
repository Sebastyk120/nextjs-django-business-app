import io
import math
import os
import time
from datetime import datetime, timedelta, date

import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test, login_required
from django.db import transaction
from django.db.models import Sum
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden, FileResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.decorators import method_decorator
from django.views import View
from django.views.generic import TemplateView
from django.views.generic.edit import CreateView, UpdateView
from django_tables2 import SingleTableView
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from xhtml2pdf import pisa

from mysite import settings
from .forms import SearchForm, PedidoForm, EditarPedidoForm, EliminarPedidoForm, DetallePedidoForm, \
    EliminarDetallePedidoForm, EditarPedidoExportadorForm, EditarDetallePedidoForm, EditarReferenciaForm, \
    EditarPedidoSeguimientoForm, FiltroSemanaExportadoraForm, SearchFormReferencias, EditarPedidoFormDos, \
    EditarPedidoFormCartera, EditarPedidoFormUtilidades, EditarDetallePedidoDosForm, EditarDetallePedidoTresForm, \
    ExportSearchForm, ExportSearchFormSeguimientos
from .models import Pedido, DetallePedido, Referencias, AutorizacionCancelacion, Presentacion, Exportador, Cliente, \
    Intermediario
from .resources import obtener_datos_con_totales_cliente, crear_archivo_excel_cliente, \
    crear_archivo_excel_enviar_cliente, obtener_datos_con_totales_enviar_cliente
from .tables import PedidoTable, DetallePedidoTable, PedidoExportadorTable, CarteraPedidoTable, UtilidadPedidoTable, \
    ResumenPedidoTable, ReferenciasTable, SeguimienosTable, SeguimienosResumenTable


# -----------Funcion para permisos por grupo ---------------------
def es_miembro_del_grupo(nombre_grupo):
    def es_miembro(user):
        return user.groups.filter(name=nombre_grupo).exists()

    return es_miembro


@login_required
def redirect_based_on_group_pedidos(request):
    user = request.user
    if user.groups.filter(name='Heavens').exists():
        return redirect('pedido_list_general')
    elif user.groups.filter(name='Fieldex').exists():
        return redirect('pedido_list_fieldex')
    elif user.groups.filter(name='Etnico').exists():
        return redirect('pedido_list_etnico')
    elif user.groups.filter(name='Juan_Matas').exists():
        return redirect('pedido_list_juan')
    elif user.groups.filter(name='CI_Dorado').exists():
        return redirect('pedido_list_ci_dorado')
    else:
        # Redirigir a una vista por defecto si el usuario no pertenece a ninguno de los grupos
        return redirect('home')


@login_required
def redirect_based_on_group_cartera(request):
    user = request.user
    if user.groups.filter(name='Heavens').exists():
        return redirect('cartera_list_heavens')
    elif user.groups.filter(name='Fieldex').exists():
        return redirect('cartera_list_fieldex')
    elif user.groups.filter(name='Etnico').exists():
        return redirect('cartera_list_etnico')
    elif user.groups.filter(name='Juan_Matas').exists():
        return redirect('cartera_list_juan')
    elif user.groups.filter(name='CI_Dorado').exists():
        return redirect('cartera_list_ci_dorado')
    else:
        # Redirigir a una vista por defecto si el usuario no pertenece a ninguno de los grupos
        return redirect('home')


# ----------------- Resumen Exportaciones Table View -------------------------------------

@method_decorator(login_required, name='dispatch')
class ResumenPedidoListView(SingleTableView):
    model = DetallePedido
    table_class = ResumenPedidoTable
    template_name = 'resumen_pedido.html'

    def dispatch(self, request, *args, **kwargs):
        pedido_id = self.kwargs.get('pedido_id')
        pedido = get_object_or_404(Pedido, pk=pedido_id)
        # Comprueba si el usuario pertenece al grupo requerido
        if not request.user.groups.filter(name=pedido.exportadora.nombre).exists():
            return HttpResponseForbidden("No tienes permiso para ver estos detalles del pedido")

        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        pedido_id = self.kwargs.get('pedido_id')
        return DetallePedido.objects.filter(pedido__id=pedido_id)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pedido_id = self.kwargs.get('pedido_id')
        context['pedido'] = get_object_or_404(Pedido, pk=pedido_id)
        queryset = self.get_queryset()

        # Calcular el total de cajas solicitadas
        total_cajas_solicitadas = self.get_queryset().aggregate(Sum('cajas_solicitadas'))['cajas_solicitadas__sum']
        total_peso_bruto = sum(obj.calcular_peso_bruto() for obj in queryset)
        total_piezas = math.ceil(float(sum(obj.calcular_no_piezas() for obj in queryset)))
        context['total_cajas_solicitadas'] = total_cajas_solicitadas
        context['total_peso_bruto'] = total_peso_bruto
        context['total_piezas'] = total_piezas

        return context


# ------------------ Exportacion de Detalles de Pedidos Excel General -------------------------------------
@login_required
@user_passes_test(es_miembro_del_grupo('Heavens'), login_url='home')
def exportar_detalles_pedidos_excel(request):
    # Crear un buffer en memoria
    output = io.BytesIO()

    # Crear un libro de trabajo (no usar write_only para poder modificar propiedades)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Detalles De Pedidos General'

    # Definir estilos para los encabezados
    header_font = Font(name='Arial', bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Estilos para filas alternas
    even_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    # Estilos para totales
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # Definir las columnas
    columnas = [
        'Pedido', 'F Entrega', 'Exportador', 'AWB', 'Cliente', 'Fruta', 'Presentacion', 'Cajas Solicitadas',
        'Peso Presentacion', 'kilos', 'Cajas Enviadas', 'Kilos Enviados', 'Diferencia', 'Tipo Caja',
        'Referencia', 'Stiker', 'Lleva Contenedor', 'Ref Contenedor', 'Cant Contenedor', 'Tarifa utilidad',
        'Tarifa Recuperacion',
        'Valor x Caja USD', 'Valor X Producto', 'No Cajas NC', 'Valor NC', 'Afecta utilidad',
        'Valor Total utilidad Producto', 'Total Recuperacion X Producto', 'Precio Proforma', 'Observaciones'
    ]

    # Anchos aproximados para cada columna
    column_widths = {
        'Pedido': 10, 'F Entrega': 12, 'Exportador': 20, 'AWB': 15, 'Cliente': 20,
        'Fruta': 15, 'Presentacion': 15, 'Cajas Solicitadas': 15, 'Peso Presentacion': 15,
        'kilos': 10, 'Cajas Enviadas': 15, 'Kilos Enviados': 15, 'Diferencia': 12,
        'Tipo Caja': 15, 'Referencia': 15, 'Stiker': 15, 'Lleva Contenedor': 15,
        'Ref Contenedor': 15, 'Cant Contenedor': 15, 'Tarifa utilidad': 15,
        'Tarifa Recuperacion': 15, 'Valor x Caja USD': 15, 'Valor X Producto': 15,
        'No Cajas NC': 12, 'Valor NC': 12, 'Afecta utilidad': 15,
        'Valor Total utilidad Producto': 22, 'Total Recuperacion X Producto': 22,
        'Precio Proforma': 15, 'Observaciones': 30
    }

    # Identificar columnas numéricas para aplicar formato después
    numeric_columns = [
        'Cajas Solicitadas', 'Peso Presentacion', 'kilos', 'Cajas Enviadas', 'Kilos Enviados',
        'Diferencia', 'Cant Contenedor', 'Tarifa utilidad', 'Tarifa Recuperacion',
        'Valor x Caja USD', 'Valor X Producto', 'No Cajas NC', 'Valor NC',
        'Valor Total utilidad Producto', 'Total Recuperacion X Producto', 'Precio Proforma'
    ]
    numeric_indices = [columnas.index(col) for col in numeric_columns]

    # Columnas monetarias para formato de moneda
    currency_columns = [
        'Tarifa utilidad', 'Tarifa Recuperacion', 'Valor x Caja USD', 'Valor X Producto',
        'Valor NC', 'Valor Total utilidad Producto', 'Total Recuperacion X Producto', 'Precio Proforma'
    ]
    currency_indices = [columnas.index(col) for col in currency_columns]

    # Agregar encabezados con estilos
    for col_idx, column_title in enumerate(columnas, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

        # Ajustar ancho de columnas
        worksheet.column_dimensions[get_column_letter(col_idx)].width = column_widths.get(column_title, 15)

    # Obtener filtros desde el request
    numero_pedido_inicial = request.POST.get('numero_pedido_inicial')
    numero_pedido_final = request.POST.get('numero_pedido_final')

    try:
        if numero_pedido_inicial and numero_pedido_final:
            numero_pedido_inicial = int(numero_pedido_inicial)
            numero_pedido_final = int(numero_pedido_final)
            queryset = DetallePedido.objects.select_related(
                'pedido__exportadora',
                'pedido__cliente',
                'fruta',
                'presentacion',
                'tipo_caja',
                'referencia',
                # Agrega otros campos relacionados según sea necesario
            ).filter(pedido__pk__range=(numero_pedido_inicial, numero_pedido_final))
        else:
            queryset = DetallePedido.objects.select_related(
                'pedido__exportadora',
                'pedido__cliente',
                'fruta',
                'presentacion',
                'tipo_caja',
                'referencia',
                # Agrega otros campos relacionados según sea necesario
            ).all()

        # Variables para totales
        totals = {idx: 0 for idx in numeric_indices}
        row_idx = 2  # Comenzamos en la fila 2 después del encabezado

        # Iterar sobre el queryset sin almacenar en caché
        for detalle in queryset.iterator():
            fila = [
                detalle.pedido.pk,
                detalle.pedido.fecha_entrega,
                detalle.pedido.exportadora.nombre,
                detalle.pedido.awb,
                detalle.pedido.cliente.nombre,
                detalle.fruta.nombre,
                detalle.presentacion.nombre,
                detalle.cajas_solicitadas,
                detalle.presentacion_peso,
                detalle.kilos,
                detalle.cajas_enviadas,
                detalle.kilos_enviados,
                detalle.diferencia,
                detalle.tipo_caja.nombre,
                detalle.referencia.nombre,
                detalle.stickers,
                detalle.lleva_contenedor,
                detalle.referencia_contenedor,
                detalle.cantidad_contenedores,
                detalle.tarifa_utilidad,
                detalle.tarifa_recuperacion,
                detalle.valor_x_caja_usd,
                detalle.valor_x_producto,
                detalle.no_cajas_nc,
                detalle.valor_nota_credito_usd,
                detalle.afecta_utilidad,
                detalle.valor_total_utilidad_x_producto,
                detalle.valor_total_recuperacion_x_producto,
                detalle.precio_proforma,
                detalle.observaciones,
            ]

            # Sumar a los totales
            for idx in numeric_indices:
                if isinstance(fila[idx], (int, float)) and fila[idx] is not None:
                    totals[idx] += fila[idx]

            # Aplicar estilo a filas alternas
            for col_idx, value in enumerate(fila, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Side(style='thin')

                # Formato de fecha
                if col_idx == 2 and isinstance(value, datetime):  # F Entrega
                    cell.number_format = 'DD/MM/YYYY'

                # Formato de números
                if col_idx - 1 in numeric_indices:
                    cell.number_format = '#,##0.00'

                # Formato de moneda
                if col_idx - 1 in currency_indices:
                    cell.number_format = '"$"#,##0.00'

                # Color de fondo para filas pares
                if row_idx % 2 == 0:
                    cell.fill = even_fill

            row_idx += 1

        # Agregar fila de totales
        row_idx += 1
        for col_idx in range(1, len(columnas) + 1):
            if col_idx - 1 in numeric_indices:
                cell = worksheet.cell(row=row_idx, column=col_idx, value=totals[col_idx - 1])
                cell.font = total_font
                cell.fill = total_fill

                # Aplicar formato de número
                cell.number_format = '#,##0.00'

                # Aplicar formato de moneda si corresponde
                if col_idx - 1 in currency_indices:
                    cell.number_format = '"$"#,##0.00'
            elif col_idx == 1:
                cell = worksheet.cell(row=row_idx, column=col_idx, value="TOTALES")
                cell.font = total_font
                cell.fill = total_fill
            else:
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = total_fill

        # Aplicar autofilter a todo el rango de datos
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}{row_idx - 1}"

        # Congelar la primera fila para que sea visible al desplazarse
        worksheet.freeze_panes = "A2"

    except ValueError:
        return HttpResponse("Número de pedido inválido", status=400)

    # Guardar el libro de trabajo en el buffer
    workbook.save(output)
    output.seek(0)

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Detalles_Pedidos.xlsx"'

    return response



# -------------------------- Funciones De Exportacion Cartera General Antigua--------------------------------------------------
class ExportarCarteraClienteVistaAntiguaView(TemplateView):
    template_name = 'export_cartera_cliente.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ExportSearchForm()
        return context

    def post(self, request, *args, **kwargs):
        form = ExportSearchForm(request.POST)
        if form.is_valid():
            cliente = form.cleaned_data.get('cliente', None)
            intermediario = form.cleaned_data.get('intermediario', None)
            fecha_inicial_str = request.POST.get('fecha_inicial', None)
            fecha_final_str = request.POST.get('fecha_final', None)

            fecha_inicial = datetime.strptime(fecha_inicial_str, "%Y-%m-%d") if fecha_inicial_str else None
            fecha_final = datetime.strptime(fecha_final_str, "%Y-%m-%d") if fecha_final_str else None

            # Determinar el grupo del usuario
            grupo = None
            if es_miembro_del_grupo('Heavens')(request.user):
                grupo = 'Heavens'
            elif es_miembro_del_grupo('Etnico')(request.user):
                grupo = 'Etnico'
            elif es_miembro_del_grupo('Fieldex')(request.user):
                grupo = 'Fieldex'
            elif es_miembro_del_grupo('Juan_Matas')(request.user):
                grupo = 'Juan_Matas'
            elif es_miembro_del_grupo('CI_Dorado')(request.user):
                grupo = 'CI_Dorado'

            # Obtener los datos y los totales con el filtro de fecha, cliente, intermediario y grupo
            pedidos, totales = obtener_datos_con_totales_cliente(fecha_inicial, fecha_final, cliente, intermediario,
                                                                 grupo)

            # Crear el archivo Excel
            ruta_archivo = 'estado_cuenta_clientes.xlsx'
            crear_archivo_excel_cliente(pedidos, totales, ruta_archivo)

            # Leer el archivo y preparar la respuesta
            with open(ruta_archivo, 'rb') as archivo_excel:
                response = HttpResponse(archivo_excel.read(),
                                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="estado_cuenta_clientes.xlsx"'

            return response

        context = self.get_context_data(**kwargs)
        context['form'] = form
        return self.render_to_response(context)


# ------------------ Exportacion De Cartera Por General Dash y demas --------------------------------------------------------------
class ExportarCarteraClienteView(View):
    def get(self, request):
        cliente_id = request.GET.get('cliente', '')
        fecha_inicial_str = request.GET.get('fecha_inicial', '')
        fecha_final_str = request.GET.get('fecha_final', '')
        grupo = request.GET.get('grupo', '')

        hoy = datetime.now().date()
        primer_dia_mes = datetime(hoy.year, hoy.month, 1).date()

        # Convertir las fechas recibidas o asignar valores por defecto
        fecha_inicial = parse_date(fecha_inicial_str) or primer_dia_mes
        fecha_final = parse_date(fecha_final_str) or hoy

        # Si se envió un cliente, obtener el objeto, sino dejarlo en None (si corresponde)
        cliente = Cliente.objects.get(id=cliente_id) if cliente_id else None

        # Usar keyword arguments para clarificar lo que se envía
        data = obtener_datos_con_totales_cliente(
            fecha_inicial=fecha_inicial,
            fecha_final=fecha_final,
            cliente=cliente,
            grupo=grupo
        )
        return HttpResponse(status=204)


# ------------------  Exportacion De Cartera Por cliente  Para Enviar --------------------------------------------

class ExportarCarteraClienteEnviarView(TemplateView):
    template_name = 'export_cartera_cliente_enviar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ExportSearchForm()
        return context

    def post(self, request, *args, **kwargs):
        form = ExportSearchForm(request.POST)
        if form.is_valid():
            cliente = form.cleaned_data.get('cliente', None)
            intermediario = form.cleaned_data.get('intermediario', None)
            fecha_inicial_str = request.POST.get('fecha_inicial', None)
            fecha_final_str = request.POST.get('fecha_final', None)

            fecha_inicial = datetime.strptime(fecha_inicial_str, "%Y-%m-%d") if fecha_inicial_str else None
            fecha_final = datetime.strptime(fecha_final_str, "%Y-%m-%d") if fecha_final_str else None

            # Determinar el grupo del usuario
            grupo = None
            if es_miembro_del_grupo('Heavens')(request.user):
                grupo = 'Heavens'
            elif es_miembro_del_grupo('Etnico')(request.user):
                grupo = 'Etnico'
            elif es_miembro_del_grupo('Fieldex')(request.user):
                grupo = 'Fieldex'
            elif es_miembro_del_grupo('Juan_Matas')(request.user):
                grupo = 'Juan_Matas'

            # Obtener los datos y los totales con el filtro de fecha, cliente, intermediario y grupo
            pedidos, totales = obtener_datos_con_totales_enviar_cliente(fecha_inicial, fecha_final, cliente,
                                                                        intermediario,
                                                                        grupo)

            # Crear el archivo Excel
            ruta_archivo = 'estado_cuenta_clientes.xlsx'
            crear_archivo_excel_enviar_cliente(pedidos, totales, ruta_archivo, fecha_inicial, fecha_final, cliente, intermediario, grupo)

            # Leer el archivo y preparar la respuesta
            with open(ruta_archivo, 'rb') as archivo_excel:
                response = HttpResponse(archivo_excel.read(),
                                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="estado_cuenta_clientes.xlsx"'

            return response

        context = self.get_context_data(**kwargs)
        context['form'] = form
        return self.render_to_response(context)


# -------------------------------- Tabla De Pedidos General  ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoListView(SingleTableView):
    model = Pedido
    table_class = PedidoTable
    table_pagination = {"per_page": 18}
    template_name = 'pedido_list_general.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().prefetch_related('autorizacioncancelacion_set')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')
            estado_pedido = form.cleaned_data.get('estado_pedido')
            fecha_desde = form.cleaned_data.get('fecha_desde')
            fecha_hasta = form.cleaned_data.get('fecha_hasta')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)
            
            if estado_pedido:
                queryset = queryset.filter(estado_pedido=estado_pedido)
                
            if fecha_desde:
                queryset = queryset.filter(fecha_entrega__gte=fecha_desde)
                
            if fecha_hasta:
                queryset = queryset.filter(fecha_entrega__lte=fecha_hasta)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        context['es_autorizador'] = self.request.user.groups.filter(name='Autorizadores').exists()
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False).distinct()


# -------------------------------- Tabla De Pedidos Seguimientos ----------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class SeguimientosPedidosListView(SingleTableView):
    model = Pedido
    table_class = SeguimienosTable
    table_pagination = {"per_page": 18}
    template_name = 'seguimiento_pedido_list_general.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().prefetch_related('autorizacioncancelacion_set')
        form = self.form_class(self.request.GET)

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'cliente':
                    queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()  # No results if ID is not an integer

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ExportSearchFormSeguimientos()
        context['es_autorizador'] = self.request.user.groups.filter(name='Autorizadores').exists()
        return context


# -------------------------------- Tabla De Pedidos Seguimientos Resumen ----------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class ResumenSeguimientosPedidosListView(SingleTableView):
    model = Pedido
    table_class = SeguimienosResumenTable
    table_pagination = {"per_page": 18}
    template_name = 'resumen_seguimiento_pedido_list_general.html'
    form_class = FiltroSemanaExportadoraForm

    def get_queryset(self):
        queryset = super().get_queryset().prefetch_related('autorizacioncancelacion_set')
        form = self.form_class(self.request.GET)

        if form.is_valid():
            semana = form.cleaned_data.get('semana')
            exportadora = form.cleaned_data.get('exportadora')

            if semana:
                queryset = queryset.filter(semana=semana)
            if exportadora:
                queryset = queryset.filter(exportadora=exportadora)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET)
        context['es_autorizador'] = self.request.user.groups.filter(name='Autorizadores').exists()
        return context


# -------------------------------- Tabla De Pedidos Etnico  ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Etnico'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoEtnicoListView(SingleTableView):
    model = Pedido
    table_class = PedidoExportadorTable
    table_pagination = {"per_page": 18}
    template_name = 'pedido_list_etnico.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Etnico')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False, pedido__exportadora__nombre='Etnico').distinct()


# -------------------------------- Tabla De Pedidos Fieldex  ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoFieldexListView(SingleTableView):
    model = Pedido
    table_class = PedidoExportadorTable
    table_pagination = {"per_page": 18}
    template_name = 'pedido_list_fieldex.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Fieldex')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False, pedido__exportadora__nombre='Fieldex').distinct()


# -------------------------------- Tabla De Pedidos Juan Matas  ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoJuanListView(SingleTableView):
    model = Pedido
    table_class = PedidoExportadorTable
    table_pagination = {"per_page": 18}
    template_name = 'pedido_list_Juan.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Juan_Matas')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False, pedido__exportadora__nombre='Juan_Matas').distinct()


# -------------------------------- Tabla De Pedidos CI  ----------------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('CI_Dorado'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoCiDoradoListView(SingleTableView):
    model = Pedido
    table_class = PedidoExportadorTable
    table_pagination = {"per_page": 18}
    template_name = 'pedido_list_ci_dorado.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='CI_Dorado')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False, pedido__exportadora__nombre='CI_Dorado').distinct()

# -------------------------------  Formulario - Crear Pedido General - Modal (General) ----------------------------
@method_decorator(login_required, name='dispatch')
class PedidoCreateView(CreateView):
    model = Pedido
    form_class = PedidoForm
    template_name = 'pedido_crear.html'
    success_url = '/pedido_list_general/'

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        return form

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request,
                         f"El pedido para el cliente {form.cleaned_data['cliente']} se ha creado exitosamente.")
        return JsonResponse({'success': True})

    def form_invalid(self, form):
        return JsonResponse({'success': False, 'html': render_to_string(self.template_name, {'form': form})})


# -------------------------------  Formulario - Editar Pedido General - Modal (General) ----------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoUpdateView(UpdateView):
    model = Pedido
    form_class = EditarPedidoForm
    template_name = 'pedido_editar.html'
    success_url = '/pedido_list_general/'

    def get_object(self, queryset=None):
        pedido_id = int(self.request.POST.get('pedido_id').replace(".", "")) if self.request.POST.get(
            'pedido_id') else int(self.request.GET.get('pedido_id').replace(".", ""))
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido'] = self.object
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()

        formatted_fecha_entrega = self.object.fecha_entrega.strftime('%Y-%m-%d') if self.object.fecha_entrega else ''
        formatted_fecha_pago_utilidad = self.object.fecha_pago_utilidad.strftime(
            '%Y-%m-%d') if self.object.fecha_pago_utilidad else ''

        form = self.form_class(
            instance=self.object,
            initial={'fecha_entrega': formatted_fecha_entrega, 'fecha_pago_utilidad': formatted_fecha_pago_utilidad}
        )

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form, 'pedido': self.object}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(self.get_context_data(form=form))

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.object
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, f"El pedido ha sido editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False,
                                 'html': render_to_string(self.template_name, {'form': form, 'pedido': self.object},
                                                          request=self.request)})
        else:
            return super().form_invalid(form)


# -------------------------------  Formulario - Editar Pedido Segunda parte - Modal (General) ----------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoUpdateViewDos(UpdateView):
    model = Pedido
    form_class = EditarPedidoFormDos
    template_name = 'pedido_editar2.html'
    success_url = '/pedido_list_general/'

    def get_object(self, queryset=None):
        pedido_id = int(self.request.POST.get('pedido_id').replace(".", "")) if self.request.POST.get(
            'pedido_id') else int(self.request.GET.get('pedido_id').replace(".", ""))
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido'] = self.object
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()

        formatted_fecha_entrega = self.object.fecha_entrega.strftime('%Y-%m-%d') if self.object.fecha_entrega else ''
        formatted_fecha_pago_utilidad = self.object.fecha_pago_utilidad.strftime(
            '%Y-%m-%d') if self.object.fecha_pago_utilidad else ''

        form = self.form_class(
            instance=self.object,
            initial={'fecha_entrega': formatted_fecha_entrega, 'fecha_pago_utilidad': formatted_fecha_pago_utilidad}
        )

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form, 'pedido': self.object}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(self.get_context_data(form=form))

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.object
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, f"El pedido ha sido editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False,
                                 'html': render_to_string(self.template_name, {'form': form, 'pedido': self.object},
                                                          request=self.request)})
        else:
            return super().form_invalid(form)


# -------------------------------  Formulario - Editar Pedido Cartera - Modal (General) ----------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoUpdateViewCartera(UpdateView):
    model = Pedido
    form_class = EditarPedidoFormCartera
    template_name = 'pedido_editar_cartera.html'
    success_url = '/cartera_list_heavens'

    def get_object(self, queryset=None):
        pedido_id = int(self.request.POST.get('pedido_id').replace(".", "")) if self.request.POST.get(
            'pedido_id') else int(self.request.GET.get('pedido_id').replace(".", ""))
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido'] = self.object
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()

        formatted_fecha_entrega = self.object.fecha_entrega.strftime('%Y-%m-%d') if self.object.fecha_entrega else ''
        formatted_fecha_pago_utilidad = self.object.fecha_pago_utilidad.strftime(
            '%Y-%m-%d') if self.object.fecha_pago_utilidad else ''

        form = self.form_class(
            instance=self.object,
            initial={'fecha_entrega': formatted_fecha_entrega, 'fecha_pago_utilidad': formatted_fecha_pago_utilidad}
        )

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form, 'pedido': self.object}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(self.get_context_data(form=form))

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.object
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, f"El pedido ha sido editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False,
                                 'html': render_to_string(self.template_name, {'form': form, 'pedido': self.object},
                                                          request=self.request)})
        else:
            return super().form_invalid(form)


# --------------------------------------- Editar pedido Utilidades -------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoUpdateViewUtilidades(UpdateView):
    model = Pedido
    form_class = EditarPedidoFormUtilidades
    template_name = 'pedido_editar_utilidades.html'
    success_url = '/utilidad_list_heavens/'

    def get_object(self, queryset=None):
        pedido_id = int(self.request.POST.get('pedido_id').replace(".", "")) if self.request.POST.get(
            'pedido_id') else int(self.request.GET.get('pedido_id').replace(".", ""))
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido'] = self.object
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()

        formatted_fecha_entrega = self.object.fecha_entrega.strftime('%Y-%m-%d') if self.object.fecha_entrega else ''
        formatted_fecha_pago_utilidad = self.object.fecha_pago_utilidad.strftime(
            '%Y-%m-%d') if self.object.fecha_pago_utilidad else ''

        form = self.form_class(
            instance=self.object,
            initial={'fecha_entrega': formatted_fecha_entrega, 'fecha_pago_utilidad': formatted_fecha_pago_utilidad}
        )

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form, 'pedido': self.object}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(self.get_context_data(form=form))

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.object
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, f"El pedido ha sido editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False,
                                 'html': render_to_string(self.template_name, {'form': form, 'pedido': self.object},
                                                          request=self.request)})
        else:
            return super().form_invalid(form)


# -------------------------------  //// Formulario - Editar Pedido Por Exportador //// ----------------------------
@method_decorator(login_required, name='dispatch')
class PedidoExportadorUpdateView(UpdateView):
    model = Pedido
    form_class = EditarPedidoExportadorForm
    template_name = 'pedido_editar_exportador.html'
    success_url = '/update_items/'

    def get_object(self, queryset=None):
        pedido_id = int(self.request.POST.get('pedido_id').replace(".", "")) if self.request.POST.get(
            'pedido_id') else int(self.request.GET.get('pedido_id').replace(".", ""))
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido'] = self.object
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()

        formatted_fecha_monetizacion = self.object.fecha_monetizacion.strftime(
            '%Y-%m-%d') if self.object.fecha_monetizacion else ''
        formatted_fecha_pago = self.object.fecha_pago.strftime('%Y-%m-%d') if self.object.fecha_pago else ''

        form = self.form_class(
            instance=self.object,
            initial={'fecha_monetizacion': formatted_fecha_monetizacion, 'fecha_pago': formatted_fecha_pago}
        )

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form, 'pedido': self.object}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(self.get_context_data(form=form))

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, f"El pedido ha sido editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False,
                                 'html': render_to_string(self.template_name, {'form': form, 'pedido': self.object},
                                                          request=self.request)})
        else:
            return super().form_invalid(form)


# -------------------------------  Formulario - Editar Pedido Seguimiento - Modal (Tracking)----------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoUpdateSebguimientoView(UpdateView):
    model = Pedido
    form_class = EditarPedidoSeguimientoForm
    template_name = 'pedido_editar_seguimiento.html'
    success_url = '/seguimiento_pedido_list_general/'

    def get_object(self, queryset=None):
        pedido_id = int(self.request.POST.get('pedido_id').replace(".", "")) if self.request.POST.get(
            'pedido_id') else int(self.request.GET.get('pedido_id').replace(".", ""))
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido'] = self.object
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        formatted_fecha_llegada = self.object.etd.strftime('%Y-%m-%d') if self.object.etd else ''
        formatted_etd = self.object.etd.strftime('%Y-%m-%dT%H:%M') if self.object.etd else ''
        formatted_eta = self.object.eta.strftime('%Y-%m-%dT%H:%M') if self.object.eta else ''
        formatted_eta_real = self.object.eta_real.strftime('%Y-%m-%dT%H:%M') if self.object.eta_real else ''

        form = self.form_class(
            instance=self.object,
            initial={'fecha_llegada': formatted_fecha_llegada, 'etd': formatted_etd, 'eta': formatted_eta,
                     'eta_real': formatted_eta_real})

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form, 'pedido': self.object}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(self.get_context_data(form=form))

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, f"El pedido ha sido editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False,
                                 'html': render_to_string(self.template_name, {'form': form, 'pedido': self.object},
                                                          request=self.request)})
        else:
            return super().form_invalid(form)


# -------------------------------  Formulario - Eliminar Pedido General - Modal (General) ----------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoDeleteView(UpdateView):
    model = Pedido
    form_class = EliminarPedidoForm
    template_name = 'pedido_eliminar.html'
    success_url = '/pedido_list_general/'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.object = None

    def get_object(self, queryset=None):
        pedido_id = int(self.request.POST.get('pedido_id').replace(".", ""))
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get(self, request, *args, **kwargs):
        pedido_id = int(request.GET.get('pedido_id').replace(".", ""))
        self.object = get_object_or_404(Pedido, id=pedido_id)
        formatted_fecha_solicitud = self.object.fecha_solicitud.strftime('%Y-%m-%d')
        formatted_fecha_entrega = self.object.fecha_entrega.strftime('%Y-%m-%d')
        form = self.form_class(
            instance=self.object,
            initial={'fecha_solicitud': formatted_fecha_solicitud, 'fecha_entrega': formatted_fecha_entrega}
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        pedido = form.save(commit=False)
        pedido.delete()
        messages.warning(self.request,
                         f"El pedido para el cliente {form.cleaned_data['cliente']} se ha eliminado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {'success': False, 'html': render_to_string(self.template_name, {'form': form}, request=self.request)})
        else:
            return super().form_invalid(form)


# ----------------------------------- Mostrar Detalles De Pedido General ------------------------------------------
@method_decorator(login_required, name='dispatch')
class DetallePedidoListView(SingleTableView):
    model = DetallePedido
    table_class = DetallePedidoTable
    template_name = 'pedido_detalle_list.html'

    def dispatch(self, request, *args, **kwargs):
        pedido_id = self.kwargs.get('pedido_id')
        pedido = get_object_or_404(Pedido, pk=pedido_id)

        # Comprueba si el usuario pertenece al grupo requerido
        if not request.user.groups.filter(name=pedido.exportadora.nombre).exists():
            return HttpResponseForbidden("No tienes permiso para ver estos detalles del pedido")

        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        pedido_id = self.kwargs.get('pedido_id')
        queryset = DetallePedido.objects.filter(pedido__id=pedido_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pedido_id = self.kwargs.get('pedido_id')
        context['pedido'] = get_object_or_404(Pedido, pk=pedido_id)
        return context


# --------------------------- Formulario Crear  Detalle De Pedido ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class DetallePedidoCreateView(CreateView):
    model = DetallePedido
    form_class = DetallePedidoForm
    template_name = 'detalle_pedido_crear.html'
    success_url = '/detalle_pedido_crear/'

    def get_initial(self):
        initial = super().get_initial()
        pedido_id = self.kwargs.get('pedido_id') or self.request.GET.get('pedido_id')
        if pedido_id:
            initial['pedido'] = pedido_id
        return initial

    def get_form_kwargs(self):
        kwargs = super(DetallePedidoCreateView, self).get_form_kwargs()
        kwargs['pedido_id'] = self.kwargs.get('pedido_id')
        return kwargs

    @transaction.atomic
    def form_valid(self, form):
        pedido_id = self.kwargs.get('pedido_id')
        if pedido_id:
            pedido = get_object_or_404(Pedido, pk=pedido_id)
            form.instance.pedido = pedido

        self.object = form.save()
        messages.success(self.request, f'El detalle de pedido para el pedido {pedido_id} se ha creado exitosamente.')
        return JsonResponse({'success': True})

    def form_invalid(self, form):
        errors = form.errors.as_json()
        return JsonResponse({'success': False, 'error': errors})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido_id'] = self.kwargs.get('pedido_id')
        return context


# ---------------------------- Formulario Editar Detalle De Pedido --------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class DetallePedidoUpdateView(UpdateView):
    model = DetallePedido
    form_class = EditarDetallePedidoForm
    template_name = 'detalle_pedido_editar.html'
    success_url = reverse_lazy('pedido_detalle_list')

    def get_object(self, queryset=None):
        detallepedido_id = self.request.GET.get('detallepedido_id') or self.request.POST.get('detallepedido_id')
        if not detallepedido_id:
            raise ValueError("No se proporcionó 'detallepedido_id'")
        return get_object_or_404(DetallePedido, id=int(detallepedido_id))

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        pedido_id = request.GET.get('pedido_id')
        form = self.form_class(instance=self.object, pedido_id=pedido_id)
        context = self.get_context_data(form=form)

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, context, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        pedido_id = request.POST.get('pedido_id')
        form = self.form_class(request.POST, instance=self.object, pedido_id=pedido_id)
        context = self.get_context_data(form=form)

        if form.is_valid():
            return self.form_valid(form, pedido_id)
        else:
            return self.form_invalid(form, context)

    @transaction.atomic
    def form_valid(self, form, pedido_id):
        self.object = form.save()
        messages.success(self.request,
                         f"El detalle para el pedido {pedido_id} se ha editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form, context):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, context, request=self.request)
            return JsonResponse({'success': False, 'form_html': form_html})
        else:
            return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido_id'] = self.request.GET.get('pedido_id') or self.request.POST.get('pedido_id')
        return context


# --------------------------- Formulario Editar Momento 2 Detalle De Pedido ------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class DetallePedidoUpdateDosView(UpdateView):
    model = DetallePedido
    form_class = EditarDetallePedidoDosForm
    template_name = 'detalle_pedido_editar2.html'
    success_url = reverse_lazy('pedido_detalle_list')

    def get_object(self, queryset=None):
        detallepedido_id = self.request.GET.get('detallepedido_id') or self.request.POST.get('detallepedido_id')
        if not detallepedido_id:
            raise ValueError("No se proporcionó 'detallepedido_id'")
        return get_object_or_404(DetallePedido, id=int(detallepedido_id))

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        pedido_id = request.GET.get('pedido_id')
        form = self.form_class(instance=self.object, pedido_id=pedido_id)
        context = self.get_context_data(form=form)

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, context, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        pedido_id = request.POST.get('pedido_id')
        form = self.form_class(request.POST, instance=self.object, pedido_id=pedido_id)
        context = self.get_context_data(form=form)

        if form.is_valid():
            return self.form_valid(form, pedido_id)
        else:
            return self.form_invalid(form, context)

    @transaction.atomic
    def form_valid(self, form, pedido_id):
        self.object = form.save()
        messages.success(self.request,
                         f"El detalle para el pedido {pedido_id} se ha editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form, context):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, context, request=self.request)
            return JsonResponse({'success': False, 'form_html': form_html})
        else:
            return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido_id'] = self.request.GET.get('pedido_id') or self.request.POST.get('pedido_id')
        return context


# --------------------------- Formulario Editar Momento 3 Detalle De Pedido ------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class DetallePedidoUpdateTresView(UpdateView):
    model = DetallePedido
    form_class = EditarDetallePedidoTresForm
    template_name = 'detalle_pedido_editar3.html'
    success_url = reverse_lazy('pedido_detalle_list')

    def get_object(self, queryset=None):
        detallepedido_id = self.request.GET.get('detallepedido_id') or self.request.POST.get('detallepedido_id')
        if not detallepedido_id:
            raise ValueError("No se proporcionó 'detallepedido_id'")
        return get_object_or_404(DetallePedido, id=int(detallepedido_id))

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        pedido_id = request.GET.get('pedido_id')
        form = self.form_class(instance=self.object, pedido_id=pedido_id)
        context = self.get_context_data(form=form)

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, context, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        pedido_id = request.POST.get('pedido_id')
        form = self.form_class(request.POST, instance=self.object, pedido_id=pedido_id)
        context = self.get_context_data(form=form)

        if form.is_valid():
            return self.form_valid(form, pedido_id)
        else:
            return self.form_invalid(form, context)

    @transaction.atomic
    def form_valid(self, form, pedido_id):
        self.object = form.save()
        messages.success(self.request,
                         f"El detalle para el pedido {pedido_id} se ha editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form, context):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, context, request=self.request)
            return JsonResponse({'success': False, 'form_html': form_html})
        else:
            return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido_id'] = self.request.GET.get('pedido_id') or self.request.POST.get('pedido_id')
        return context


# ---------------------------- Formulario Eliminar Detalle De Pedido --------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class DetallePedidoDeleteiew(UpdateView):
    model = DetallePedido
    form_class = EliminarDetallePedidoForm
    template_name = 'detalle_pedido_eliminar.html'
    success_url = ''

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.object = None

    def get_object(self, queryset=None):
        detallepedido_id = int(self.request.POST.get('detallepedido_id').replace(".", ""))
        detallepedido = get_object_or_404(DetallePedido, id=detallepedido_id)
        return detallepedido

    def get(self, request, *args, **kwargs):
        detallepedido_id = int(request.GET.get('detallepedido_id').replace(".", ""))
        self.object = get_object_or_404(DetallePedido, id=detallepedido_id)
        form = self.form_class(
            instance=self.object,
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return super().get(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.object
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        detallepedido = form.save(commit=False)
        detallepedido.delete()
        messages.warning(self.request,
                         f"El detalle de {form.cleaned_data['pedido']} se ha eliminado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {'success': False, 'html': render_to_string(self.template_name, {'form': form}, request=self.request)})
        else:
            return super().form_invalid(form)
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido_id'] = self.request.GET.get('pedido_id') or self.request.POST.get('pedido_id')
        return context


# ------------------------- CARTERA General /// Table mostrar cartera de pedidos Heavens ///  -------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class CarteraHeavensListView(SingleTableView):
    model = Pedido
    table_class = CarteraPedidoTable
    table_pagination = {"per_page": 18}
    template_name = 'cartera_list_heavens.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset()
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        # Añadir el formulario de exportación al contexto
        context['export_form'] = ExportSearchForm()
        return context
    
    def get_clientes_con_pedidos(self):
            return Cliente.objects.filter(pedido__isnull=False).distinct()


# ------------------------- CARTERA /// Table mostrar cartera de pedidos Etnico ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Etnico'), login_url=reverse_lazy('home')), name='dispatch')
class CarteraEtnicoListView(SingleTableView):
    model = Pedido
    table_class = CarteraPedidoTable
    table_pagination = {"per_page": 18}
    template_name = 'cartera_list_etnico.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Etnico')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False, pedido__exportadora__nombre='Etnico').distinct()


# ------------------------- CARTERA /// Table mostrar cartera de pedidos Fieldex ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url=reverse_lazy('home')), name='dispatch')
class CarteraFieldexListView(SingleTableView):
    model = Pedido
    table_class = CarteraPedidoTable
    table_pagination = {"per_page": 18}
    template_name = 'cartera_list_fieldex.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Fieldex')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False, pedido__exportadora__nombre='Fieldex').distinct()


# ------------------------- CARTERA /// Table mostrar cartera de pedidos Juan Matas ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url=reverse_lazy('home')), name='dispatch')
class CarteraJuanListView(SingleTableView):
    model = Pedido
    table_class = CarteraPedidoTable
    table_pagination = {"per_page": 18}
    template_name = 'cartera_list_juan.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Juan_Matas')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False, pedido__exportadora__nombre='Juan_Matas').distinct()

# -------------------- Cartera /// Table mostrar cartera de pedidos Ci Dorado ///  -------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('CI_Dorado'), login_url=reverse_lazy('home')), name='dispatch')
class CarteraCiDoradoListView(SingleTableView):
    model = Pedido
    table_class = CarteraPedidoTable
    table_pagination = {"per_page": 18}
    template_name = 'cartera_list_ci_dorado.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='CI_Dorado')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False, pedido__exportadora__nombre='CI_Dorado').distinct()


# ------------------------- Utilidades GENERAL /// Table mostrar Utilidades de pedidos Heavens ///  ------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class UtilidadHeavensListView(SingleTableView):
    model = Pedido
    table_class = UtilidadPedidoTable
    table_pagination = {"per_page": 18}
    template_name = 'utilidad_list_heavens.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset()
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        # Añadir el formulario de exportación al contexto
        context['export_form'] = ExportSearchForm()
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False).distinct()


# ------------------------- Utilidades /// Table mostrar Utilidades de pedidos Etnico ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Etnico'), login_url=reverse_lazy('home')), name='dispatch')
class UtilidadEtnicoListView(SingleTableView):
    model = Pedido
    table_class = UtilidadPedidoTable
    table_pagination = {"per_page": 18}
    template_name = 'utilidad_list_etnico.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Etnico')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False, pedido__exportadora__nombre='Etnico').distinct()


# ------------------------- Utilidades /// Table mostrar Utilidades de pedidos Fieldex ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url=reverse_lazy('home')), name='dispatch')
class UtilidadFiedexListView(SingleTableView):
    model = Pedido
    table_class = UtilidadPedidoTable
    table_pagination = {"per_page": 18}
    template_name = 'utilidad_list_fieldex.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Fieldex')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False, pedido__exportadora__nombre='Fieldex').distinct()


# ------------------------- Utilidades /// Table mostrar Utilidades de pedidos Juan Matas // -----------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url=reverse_lazy('home')), name='dispatch')
class UtilidadJuanListView(SingleTableView):
    model = Pedido
    table_class = UtilidadPedidoTable
    table_pagination = {"per_page": 18}
    template_name = 'utilidad_list_juan.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Juan_Matas')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False, pedido__exportadora__nombre='Juan_Matas').distinct()

# --------------------------------- Utilidades CI Dorado ----------------------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('CI_Dorado'), login_url=reverse_lazy('home')), name='dispatch')
class UtilidadCiDoradoListView(SingleTableView):
    model = Pedido
    table_class = UtilidadPedidoTable
    table_pagination = {"per_page": 18}
    template_name = 'utilidad_list_ci_dorado.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='CI_Dorado')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False, pedido__exportadora__nombre='CI_Dorado').distinct()

# --------------------------------- Referencias Table Etnico----------------------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Etnico'), login_url=reverse_lazy('home')), name='dispatch')
class ReferenciasEtnicoListView(SingleTableView):
    model = Referencias
    table_class = ReferenciasTable
    template_name = 'referencia_list_etnico.html'
    form_class = SearchFormReferencias

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportador__nombre='Etnico')

        form = self.form_class(self.request.GET)
        if form.is_valid():
            item_busqueda = form.cleaned_data.get('item_busqueda')
            if item_busqueda:
                queryset = queryset.filter(nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET)
        return context


# --------------------------------- Referencias Table Fieldex----------------------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url=reverse_lazy('home')), name='dispatch')
class ReferenciasFieldexListView(SingleTableView):
    model = Referencias
    table_class = ReferenciasTable
    template_name = 'referencia_list_fieldex.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportador__nombre='Fieldex')

        form = self.form_class(self.request.GET)
        if form.is_valid():
            item_busqueda = form.cleaned_data.get('item_busqueda')
            if item_busqueda:
                queryset = queryset.filter(nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET)
        return context


# --------------------------------- Referencias Table Juan_Matas ------------------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url=reverse_lazy('home')), name='dispatch')
class ReferenciasjuanListView(SingleTableView):
    model = Referencias
    table_class = ReferenciasTable
    template_name = 'referencia_list_juan.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportador__nombre='Juan_Matas')

        form = self.form_class(self.request.GET)
        if form.is_valid():
            item_busqueda = form.cleaned_data.get('item_busqueda')
            if item_busqueda:
                queryset = queryset.filter(nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET)
        return context


# --------------------------------- Referencias Table Ci Dorado ------------------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('CI_Dorado'), login_url=reverse_lazy('home')), name='dispatch')
class ReferenciasCiDoradoListView(SingleTableView):
    model = Referencias
    table_class = ReferenciasTable
    template_name = 'referencia_list_ci_dorado.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportador__nombre='CI_Dorado')

        form = self.form_class(self.request.GET)
        if form.is_valid():
            item_busqueda = form.cleaned_data.get('item_busqueda')
            if item_busqueda:
                queryset = queryset.filter(nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET)
        return context


# ---------------------------- Formulario Editar Referencia -----------------------------------------------------------
@method_decorator(login_required, name='dispatch')
class ReferenciaUpdateView(UpdateView):
    model = Referencias
    form_class = EditarReferenciaForm
    template_name = 'referencia_editar.html'
    success_url = reverse_lazy('referencia_editar')

    def get_object(self, queryset=None):
        referencia_id = int(self.request.POST.get('referencia_id').replace(".", ""))
        referencia = get_object_or_404(Referencias, id=referencia_id)
        return referencia

    def get(self, request, *args, **kwargs):
        referencia_id = int(request.GET.get('referencia_id').replace(".", ""))
        self.object = get_object_or_404(Referencias, id=referencia_id)
        form = self.form_class(instance=self.object)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, f'La referencia {self.object.nombre} se ha editado exitosamente.')
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {'success': False, 'html': render_to_string(self.template_name, {'form': form}, request=self.request)})
        else:
            return super().form_invalid(form)


# ------------------------------Exportar Referencias a Excel-----------------------------------------------------------

# ---------------------------------- Funcion que exporta los Referencias a Excel --------------------------------------
@login_required
@user_passes_test(lambda u: any(
    es_miembro_del_grupo(grupo)(u) for grupo in ['Heavens', 'Etnico', 'Fieldex', 'Juan_Matas', 'Ci_Dorado']),
                  login_url='home')
def exportar_referencias_excel(request):
    # Determinar a qué grupo pertenece el usuario
    grupo = None
    if es_miembro_del_grupo('Heavens')(request.user):
        grupo = 'Heavens'
    elif es_miembro_del_grupo('Etnico')(request.user):
        grupo = 'Etnico'
    elif es_miembro_del_grupo('Fieldex')(request.user):
        grupo = 'Fieldex'
    elif es_miembro_del_grupo('Juan_Matas')(request.user):
        grupo = 'Juan_Matas'
    elif es_miembro_del_grupo('Ci_Dorado')(request.user):
        grupo = 'Ci_Dorado'

    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active

    # Personalizar el título de la hoja según el grupo
    if (grupo == 'Heavens'):
        worksheet.title = 'Referencias Todas'
    else:
        worksheet.title = f'Referencias {grupo}' if grupo else 'Referencias'

    font = Font(bold=True)
    fill = PatternFill(start_color="fffaac", end_color="fffaac", fill_type="solid")

    # Encabezados
    columns = ['Referencia', 'Referencia Nueva', 'Contenedor', 'Cantidad Contendedor', 'Precio', 'Exportador']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    # Filtrar referencias según el grupo del usuario
    if grupo == 'Heavens':
        # Si es Heavens, muestra TODAS las referencias (de todas las exportadoras)
        queryset = Referencias.objects.all()
    elif grupo:
        # Para otros grupos, filtrar solo las referencias de su exportadora
        queryset = Referencias.objects.filter(exportador__nombre=grupo)
    else:
        # Si no tiene grupo específico (no debería ocurrir por el decorador)
        queryset = Referencias.objects.none()

    # Agregar datos al libro de trabajo
    for row_num, referencia in enumerate(queryset, start=2):
        contenedor_nombre = referencia.contenedor.nombre if referencia.contenedor else 'Sin Contenedor'
        row = [
            referencia.nombre,
            referencia.referencia_nueva,
            contenedor_nombre,
            referencia.cant_contenedor,
            referencia.precio,
            referencia.exportador.nombre,
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Personalizar el nombre del archivo según el grupo
    if grupo == 'Heavens':
        filename = "referencias_todas.xlsx"
    else:
        filename = f"referencias_{grupo.lower() if grupo else 'general'}.xlsx"

    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

# ----------------------------- Actualizar los días de Vencimiento ----------------------------------------------------

def actualizar_dias_de_vencimiento_todos(request):
    batch_size = 150  # Tamaño del lote para evitar consumo excesivo de memoria
    pedidos = Pedido.objects.all().iterator(chunk_size=batch_size)
    pedidos_para_actualizar = []
    hoy = timezone.now().date()

    for pedido in pedidos:
        if pedido.fecha_pago is not None:
            pedido.dias_de_vencimiento = 0
        else:
            if isinstance(pedido.fecha_entrega, datetime):
                fecha_entrega = pedido.fecha_entrega.date()
            elif isinstance(pedido.fecha_entrega, date):
                fecha_entrega = pedido.fecha_entrega
            else:
                # Opcional: Manejar casos inesperados sin detener el proceso completo
                pedido.dias_de_vencimiento = None
                pedidos_para_actualizar.append(pedido)
                continue

            fecha_entrega += timedelta(days=pedido.dias_cartera)
            pedido.dias_de_vencimiento = (hoy - fecha_entrega).days

        pedidos_para_actualizar.append(pedido)

        # Actualizar en bloque cada 'batch_size' Pedidos
        if len(pedidos_para_actualizar) >= batch_size:
            with transaction.atomic():
                Pedido.objects.bulk_update(
                    pedidos_para_actualizar,
                    ['dias_de_vencimiento']
                )
            pedidos_para_actualizar = []

    # Actualizar cualquier Pedido restante
    if pedidos_para_actualizar:
        with transaction.atomic():
            Pedido.objects.bulk_update(
                pedidos_para_actualizar,
                ['dias_de_vencimiento']
            )

    messages.success(request, "Todos los pedidos se han actualizado correctamente.")
    return redirect('pedido_list_general')


# ------------------------ Vista para actualizar la TRM Banco de la republica -----------------------------------------
def actualizar_tasas(request):
    pedidos = Pedido.objects.order_by('-id')[:100]
    for pedido in pedidos:
        pedido.actualizar_tasa_representativa()
    messages.success(request, 'Se Actualizaron Las Tasas Con Banco De La Republica Correctamente')
    return redirect('pedido_list_general')


# --------------------------------- Funciones Etnico  ---------------------------------------------------------
def actualizar_dias_de_vencimiento_etnico(request):
    batch_size = 150  # Tamaño del lote para evitar consumo excesivo de memoria
    pedidos = Pedido.objects.all().iterator(chunk_size=batch_size)
    pedidos_para_actualizar = []
    hoy = timezone.now().date()

    for pedido in pedidos:
        if pedido.fecha_pago is not None:
            pedido.dias_de_vencimiento = 0
        else:
            if isinstance(pedido.fecha_entrega, datetime):
                fecha_entrega = pedido.fecha_entrega.date()
            elif isinstance(pedido.fecha_entrega, date):
                fecha_entrega = pedido.fecha_entrega
            else:
                # Opcional: Manejar casos inesperados sin detener el proceso completo
                pedido.dias_de_vencimiento = None
                pedidos_para_actualizar.append(pedido)
                continue

            fecha_entrega += timedelta(days=pedido.dias_cartera)
            pedido.dias_de_vencimiento = (hoy - fecha_entrega).days

        pedidos_para_actualizar.append(pedido)

        # Actualizar en bloque cada 'batch_size' Pedidos
        if len(pedidos_para_actualizar) >= batch_size:
            with transaction.atomic():
                Pedido.objects.bulk_update(
                    pedidos_para_actualizar,
                    ['dias_de_vencimiento']
                )
            pedidos_para_actualizar = []

    # Actualizar cualquier Pedido restante
    if pedidos_para_actualizar:
        with transaction.atomic():
            Pedido.objects.bulk_update(
                pedidos_para_actualizar,
                ['dias_de_vencimiento']
            )

    messages.success(request, "Todos los pedidos se han actualizado correctamente.")
    return redirect('pedido_list_etnico')


def actualizar_tasas_etnico(request):
    pedidos = Pedido.objects.order_by('-id')[:100]
    for pedido in pedidos:
        pedido.actualizar_tasa_representativa()
    messages.success(request, 'Se Actualizo La TRM Con Banco De La Republica Correctamente')
    return redirect('pedido_list_etnico')


# --------------------------------- Funciones Fieldex  ---------------------------------------------------------
def actualizar_dias_de_vencimiento_fieldex(request):
    batch_size = 150  # Tamaño del lote para evitar consumo excesivo de memoria
    pedidos = Pedido.objects.all().iterator(chunk_size=batch_size)
    pedidos_para_actualizar = []
    hoy = timezone.now().date()

    for pedido in pedidos:
        if pedido.fecha_pago is not None:
            pedido.dias_de_vencimiento = 0
        else:
            if isinstance(pedido.fecha_entrega, datetime):
                fecha_entrega = pedido.fecha_entrega.date()
            elif isinstance(pedido.fecha_entrega, date):
                fecha_entrega = pedido.fecha_entrega
            else:
                # Opcional: Manejar casos inesperados sin detener el proceso completo
                pedido.dias_de_vencimiento = None
                pedidos_para_actualizar.append(pedido)
                continue

            fecha_entrega += timedelta(days=pedido.dias_cartera)
            pedido.dias_de_vencimiento = (hoy - fecha_entrega).days

        pedidos_para_actualizar.append(pedido)

        # Actualizar en bloque cada 'batch_size' Pedidos
        if len(pedidos_para_actualizar) >= batch_size:
            with transaction.atomic():
                Pedido.objects.bulk_update(
                    pedidos_para_actualizar,
                    ['dias_de_vencimiento']
                )
            pedidos_para_actualizar = []

    # Actualizar cualquier Pedido restante
    if pedidos_para_actualizar:
        with transaction.atomic():
            Pedido.objects.bulk_update(
                pedidos_para_actualizar,
                ['dias_de_vencimiento']
            )

    messages.success(request, "Todos los pedidos se han actualizado correctamente.")
    return redirect('pedido_list_fieldex')


def actualizar_tasas_fieldex(request):
    pedidos = Pedido.objects.order_by('-id')[:100]
    for pedido in pedidos:
        pedido.actualizar_tasa_representativa()
    messages.success(request, 'Se Actualizo La TRM Con Banco De La Republica Correctamente')
    return redirect('pedido_list_fieldex')


# --------------------------------- Funciones Juan_matas  ---------------------------------------------------------
def actualizar_dias_de_vencimiento_juan(request):
    batch_size = 150  # Tamaño del lote para evitar consumo excesivo de memoria
    pedidos = Pedido.objects.all().iterator(chunk_size=batch_size)
    pedidos_para_actualizar = []
    hoy = timezone.now().date()

    for pedido in pedidos:
        if pedido.fecha_pago is not None:
            pedido.dias_de_vencimiento = 0
        else:
            if isinstance(pedido.fecha_entrega, datetime):
                fecha_entrega = pedido.fecha_entrega.date()
            elif isinstance(pedido.fecha_entrega, date):
                fecha_entrega = pedido.fecha_entrega
            else:
                # Opcional: Manejar casos inesperados sin detener el proceso completo
                pedido.dias_de_vencimiento = None
                pedidos_para_actualizar.append(pedido)
                continue

            fecha_entrega += timedelta(days=pedido.dias_cartera)
            pedido.dias_de_vencimiento = (hoy - fecha_entrega).days

        pedidos_para_actualizar.append(pedido)

        # Actualizar en bloque cada 'batch_size' Pedidos
        if len(pedidos_para_actualizar) >= batch_size:
            with transaction.atomic():
                Pedido.objects.bulk_update(
                    pedidos_para_actualizar,
                    ['dias_de_vencimiento']
                )
            pedidos_para_actualizar = []

    # Actualizar cualquier Pedido restante
    if pedidos_para_actualizar:
        with transaction.atomic():
            Pedido.objects.bulk_update(
                pedidos_para_actualizar,
                ['dias_de_vencimiento']
            )

    messages.success(request, "Todos los pedidos se han actualizado correctamente.")
    return redirect('pedido_list_juan')


def actualizar_tasas_juan(request):
    pedidos = Pedido.objects.order_by('-id')[:100]
    for pedido in pedidos:
        pedido.actualizar_tasa_representativa()
    messages.success(request, 'Se Actualizo La TRM Con Banco De La Republica Correctamente')
    return redirect('pedido_list_juan')


# --------------------------------- Funciones CI Dorado  ---------------------------------------------------------
def actualizar_dias_de_vencimiento_ci_dorado(request):
    batch_size = 150  # Tamaño del lote para evitar consumo excesivo de memoria
    pedidos = Pedido.objects.all().iterator(chunk_size=batch_size)
    pedidos_para_actualizar = []
    hoy = timezone.now().date()

    for pedido in pedidos:
        if pedido.fecha_pago is not None:
            pedido.dias_de_vencimiento = 0
        else:
            if isinstance(pedido.fecha_entrega, datetime):
                fecha_entrega = pedido.fecha_entrega.date()
            elif isinstance(pedido.fecha_entrega, date):
                fecha_entrega = pedido.fecha_entrega
            else:
                # Opcional: Manejar casos inesperados sin detener el proceso completo
                pedido.dias_de_vencimiento = None
                pedidos_para_actualizar.append(pedido)
                continue

            fecha_entrega += timedelta(days=pedido.dias_cartera)
            pedido.dias_de_vencimiento = (hoy - fecha_entrega).days

        pedidos_para_actualizar.append(pedido)

        # Actualizar en bloque cada 'batch_size' Pedidos
        if len(pedidos_para_actualizar) >= batch_size:
            with transaction.atomic():
                Pedido.objects.bulk_update(
                    pedidos_para_actualizar,
                    ['dias_de_vencimiento']
                )
            pedidos_para_actualizar = []

    # Actualizar cualquier Pedido restante
    if pedidos_para_actualizar:
        with transaction.atomic():
            Pedido.objects.bulk_update(
                pedidos_para_actualizar,
                ['dias_de_vencimiento']
            )

    messages.success(request, "Todos los pedidos se han actualizado correctamente.")
    return redirect('pedido_list_ci_dorado')


def actualizar_tasas_ci_dorado(request):
    pedidos = Pedido.objects.order_by('-id')[:100]
    for pedido in pedidos:
        pedido.actualizar_tasa_representativa()
    messages.success(request, 'Se Actualizo La TRM Con Banco De La Republica Correctamente')
    return redirect('pedido_list_ci_dorado')



# ----------------------------- Vista Autorización De Cancelaciones De Pedidos ---------------------------------------

# Solicitar Cancelacion Del pedido.
@login_required
def solicitar_cancelacion(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if request.method == 'POST':
        observaciones = request.POST.get('observaciones', '')
        if pedido.estado_cancelacion == 'sin_solicitud' or pedido.estado_cancelacion == 'no_autorizado':
            autorizacion = AutorizacionCancelacion.objects.create(pedido=pedido, usuario_solicitante=request.user)
            pedido.estado_cancelacion = 'pendiente'
            pedido.observaciones = observaciones
            pedido.save()
        messages.warning(request, f'Se ha enviado la solicitud de cancelación para el pedido: {pedido.id}')
        return redirect('pedido_list_general')  # Redirigir a la lista de pedidos o a donde sea necesario

    return render(request, 'solicitar_cancelacion.html', {'pedido': pedido})


# Autorizar Cancelacion del pedido
@login_required
def autorizar_cancelacion(request, autorizacion_id):
    autorizacion = get_object_or_404(AutorizacionCancelacion, id=autorizacion_id)

    if request.method == 'POST':
        observaciones = request.POST.get('observaciones', '')
        accion = request.POST.get('accion', '')
        # Verificar si el usuario actual tiene permiso para autorizar la cancelación
        if request.user.groups.filter(name='Autorizadores').exists():
            pedido = autorizacion.pedido
            if accion == 'autorizar':
                autorizacion.autorizado = True
                autorizacion.fecha_autorizacion = timezone.now()
                autorizacion.usuario_autorizador = request.user
                autorizacion.save()
                pedido.estado_cancelacion = 'autorizado'
                pedido.estado_pedido = 'Cancelado'
                pedido.estado_factura = 'Cancelada'
                pedido.awb = '-'
                pedido.descuento = 0
                pedido.dias_de_vencimiento = 0
                pedido.diferencia_por_abono = 0
                pedido.documento_cobro_utilidad = 'Pedido Cancelado'
                pedido.estado_utilidad = 'Pedido Cancelado'
                pedido.numero_factura = 'Pedido Cancelado'
                pedido.tasa_representativa_usd_diaria = 0
                pedido.total_cajas_enviadas = 0
                pedido.total_peso_bruto = 0
                pedido.total_piezas_enviadas = 0
                pedido.total_piezas_solicitadas = 0
                pedido.trm_cotizacion = 0
                pedido.trm_monetizacion = 0
                pedido.utilidad_bancaria_usd = 0
                pedido.valor_total_factura_usd = 0
                pedido.valor_total_nota_credito_usd = 0
                pedido.valor_total_utilidad_usd = 0
                pedido.valor_total_recuperacion_usd = 0
                pedido.valor_utilidad_pesos = 0
            elif accion == 'no_autorizar':
                pedido.estado_cancelacion = 'no_autorizado'
                messages.info(request, f' Se ha anulado la cancelación para el pedido: {pedido.id}')
            pedido.observaciones = observaciones
            pedido.save()

            # Eliminar todos los detalles del pedido asociado si autorizado
            if accion == 'autorizar':
                DetallePedido.objects.filter(pedido=pedido).delete()
                messages.warning(request, f' Se ha cancelado correctamente el pedido: {pedido.id}')
        return redirect('pedido_list_general')
    return render(request, 'autorizar_cancelacion.html', {'autorizacion': autorizacion})


@login_required
def filtrar_presentaciones(request):
    fruta_id = request.GET.get('fruta_id')
    pedido_id = request.GET.get('pedido_id')

    if fruta_id and pedido_id:
        presentaciones = Presentacion.objects.filter(
            clientepresentacion__cliente__pedido__id=pedido_id,
            clientepresentacion__fruta_id=fruta_id
        ).values('id', 'nombre', 'kilos')
        return JsonResponse({'presentaciones': list(presentaciones)})
    return JsonResponse({'presentaciones': []})


@login_required
def load_referencias(request):
    presentacion_id = request.GET.get('presentacion_id')
    pedido_id = request.GET.get('pedido_id')
    tipo_caja_id = request.GET.get('tipo_caja_id')
    fruta_id = request.GET.get('fruta_id')

    if presentacion_id and tipo_caja_id and fruta_id and pedido_id:
        try:
            pedido = Pedido.objects.get(id=pedido_id)
            referencias = Referencias.objects.filter(
                presentacionreferencia__presentacion_id=presentacion_id,
                presentacionreferencia__tipo_caja_id=tipo_caja_id,
                presentacionreferencia__fruta_id=fruta_id,
                exportador=pedido.exportadora
            ).distinct()
        except Pedido.DoesNotExist:
            referencias = Referencias.objects.none()
    else:
        referencias = Referencias.objects.none()

    referencias_data = [{'id': ref.id, 'nombre': ref.nombre} for ref in referencias]
    return JsonResponse({'referencias': referencias_data})


# Exportacion de Resumen a PDF
@login_required(login_url=reverse_lazy('home'))
def export_pdf_resumen_semana(request):
    # Obtener los parámetros de filtro
    semana = request.GET.get('semana')
    exportador_id = request.GET.get('exportadora')

    # Verificar que los parámetros no estén vacíos
    if not semana:
        messages.error(request, 'Debe seleccionar por lo menos una semana.')
        return redirect('resumen_seguimiento_list_heavens')

    # Filtrar los datos basados en los parámetros
    pedidos = Pedido.objects.all()
    if semana:
        pedidos = pedidos.filter(semana=semana)
    if exportador_id:
        pedidos = pedidos.filter(exportadora_id=exportador_id)
        try:
            exportador_nombre = Exportador.objects.get(id=exportador_id).nombre
        except Exportador.DoesNotExist:
            exportador_nombre = None
    else:
        exportador_nombre = None

    # Generar la tabla con los datos filtrados
    table = SeguimienosResumenTable(pedidos)
    context = {
        'table': table,
        'semana': semana,
        'exportador': exportador_nombre
    }
    html_string = render_to_string('seguimiento_resumen_pdf.html', context, request=request)

    # Función para convertir HTML a PDF
    def convert_html_to_pdf(source_html, output_filename):
        result_file = open(output_filename, "w+b")
        pisa_status = pisa.CreatePDF(source_html, dest=result_file)
        result_file.close()
        return pisa_status.err

    # Generar el PDF
    result = convert_html_to_pdf(html_string, 'output.pdf')

    if result:
        return HttpResponse("Error al generar el PDF", status=500)

    with open('output.pdf', 'rb') as pdf:
        response = HttpResponse(pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="seguimiento_resumen.pdf"'
        return response


@login_required(login_url=reverse_lazy('home'))
def exportar_pdf_resumen_pedido(request, pedido_id):
    start_time = time.time()

    # Diccionario para almacenar los tiempos de las operaciones
    tiempos = {}

    # Obtener el pedido y los detalles del pedido
    pedido = get_object_or_404(Pedido, pk=pedido_id)
    if not request.user.groups.filter(name=pedido.exportadora.nombre).exists():
        return HttpResponseForbidden("No tienes permiso para ver estos detalles del pedido")
    tiempos['verificar_permisos'] = time.time() - start_time

    # Filtrar los detalles del pedido
    detalles = DetallePedido.objects.filter(pedido=pedido).select_related('pedido')
    tiempos['filtrar_detalles'] = time.time() - start_time - tiempos['verificar_permisos']

    # Calcular totales
    total_cajas_solicitadas = detalles.aggregate(Sum('cajas_solicitadas'))['cajas_solicitadas__sum']
    total_peso_bruto = sum(detalle.calcular_peso_bruto() for detalle in detalles)
    total_piezas = math.ceil(sum(detalle.calcular_no_piezas() for detalle in detalles))
    tiempos['calcular_totales'] = time.time() - start_time - tiempos['verificar_permisos'] - tiempos['filtrar_detalles']

    # Generar la tabla con los datos filtrados
    table = ResumenPedidoTable(detalles)
    tiempos['generar_tabla'] = time.time() - start_time - tiempos['verificar_permisos'] - tiempos['filtrar_detalles'] - tiempos['calcular_totales']

    # Contexto para la plantilla
    context = {
        'pedido': pedido,
        'table': table,
        'total_cajas_solicitadas': total_cajas_solicitadas,
        'total_peso_bruto': total_peso_bruto,
        'total_piezas': total_piezas
    }

    # Renderizar la plantilla a HTML
    html_string = render_to_string('resumen_pedido_pdf.html', context, request=request)
    tiempos['renderizar_html'] = time.time() - start_time - tiempos['verificar_permisos'] - tiempos['filtrar_detalles'] - tiempos['calcular_totales'] - tiempos['generar_tabla']

    # Función para convertir HTML a PDF
    def convert_html_to_pdf(source_html, output_filename):
        result_file = open(output_filename, "w+b")
        pisa_status = pisa.CreatePDF(source_html, dest=result_file)
        result_file.close()
        return pisa_status.err

    # Generar el PDF
    result = convert_html_to_pdf(html_string, 'output.pdf')
    tiempos['generar_pdf'] = time.time() - start_time - tiempos['verificar_permisos'] - tiempos['filtrar_detalles'] - tiempos['calcular_totales'] - tiempos['generar_tabla'] - tiempos['renderizar_html']

    if result:
        return HttpResponse("Error al generar el PDF", status=500)

    with open('output.pdf', 'rb') as pdf:
        response = HttpResponse(pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="resumen_pedido_{pedido_id}.pdf"'
        tiempos['tiempo_total'] = time.time() - start_time

    # Opcional: los tiempos se pueden guardar o imprimir al final si es necesario
    return response



# -------------------- Vista para filtro de exportacion Seguimiento o Tracking -----------------------------------------


@login_required
@user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home'))
def exportar_excel_seguimiento_tracking(request):
    if request.method == 'POST':
        form = ExportSearchFormSeguimientos(request.POST)
        if form.is_valid():
            cliente = form.cleaned_data.get('cliente')
            intermediario = form.cleaned_data.get('intermediario')
            fecha_inicial = form.cleaned_data.get('fecha_inicial')
            fecha_final = form.cleaned_data.get('fecha_final')

            pedidos = Pedido.objects.all()

            if cliente:
                pedidos = pedidos.filter(cliente=cliente)
            if intermediario:
                pedidos = pedidos.filter(intermediario=intermediario)
            if fecha_inicial:
                pedidos = pedidos.filter(fecha_entrega__gte=fecha_inicial)
            if fecha_final:
                pedidos = pedidos.filter(fecha_entrega__lte=fecha_final)

            data = []
            for pedido in pedidos:
                data.append({
                    'Week': pedido.semana,
                    'Order No.': pedido.id,
                    'Request Date': pedido.fecha_solicitud.strftime('%d/%m/%Y') if pedido.fecha_solicitud else '',
                    'Exporter': pedido.exportadora.nombre if pedido.exportadora else '',
                    'Intermediary': pedido.intermediario.nombre if pedido.intermediario else '',
                    'Customer': pedido.cliente.nombre if pedido.cliente else '',
                    'Destination': pedido.destino.codigo if pedido.destino else '',
                    'Requested Boxes': pedido.total_cajas_solicitadas,
                    'Requested Pallets': pedido.total_piezas_solicitadas,
                    'Gross Weight': pedido.total_peso_bruto_solicitado,
                    'Delivery Date': pedido.fecha_entrega.strftime('%d/%m/%Y') if pedido.fecha_entrega else '',
                    'Awb': pedido.awb,
                    'Airline': pedido.aerolinea.nombre if pedido.aerolinea else '',
                    'Arrival Date': pedido.fecha_llegada.strftime('%d/%m/%Y') if pedido.fecha_llegada else '',
                    'Estimated Time of Arrival': pedido.eta.strftime('%d/%m/%Y') if pedido.eta else '',
                    'Estimated Time of Departure': pedido.etd.strftime('%d/%m/%Y') if pedido.etd else '',
                    'Cargo Agency': pedido.agencia_carga.nombre if pedido.agencia_carga else '',
                    'Products': pedido.variedades,
                    'Shipped Boxes': pedido.total_cajas_enviadas,
                    'Total Pallets Shipped': pedido.total_piezas_enviadas,
                    'Total Gross Weight Shipped	Weight': pedido.total_peso_bruto_enviado,
                    'Weight Awb': pedido.peso_awb,
                    'Kg Invoice / AWB': pedido.diferencia_peso_factura_awb,
                    'Real Final Eta': pedido.eta_real.strftime('%d/%m/%Y') if pedido.eta_real else '',
                    'Invoice': pedido.numero_factura,
                    'Termo': pedido.termo,
                    'Booking Responsible': pedido.responsable_reserva.nombre if pedido.responsable_reserva else '',
                    'Reserve Status': pedido.estatus_reserva,
                    'Document Status': pedido.estado_documentos,
                    'Order Status': pedido.estado_pedido,
                    'Comments': pedido.observaciones_tracking,
                })

            df = pd.DataFrame(data)

            wb = Workbook()
            ws = wb.active

            fill_red_soft = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        if row[df.columns.get_loc('Order Status')] == 'Cancelado':
                            cell.fill = fill_red_soft

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=seguimiento_pedidos.xlsx'
            wb.save(response)

            return response
    else:
        form = ExportSearchForm()

    return render(request, 'export_pedidos_tracking.html', {'form': form})


# -------------------- Exportacion Explicita de Excel Del resumen de exportaciones Por semana -------------------------

@login_required
@user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home'))
def exportar_excel_seguimiento_resumen(request):
    semana = request.GET.get('semana')
    exportador_id = request.GET.get('exportadora')

    # Verificar que los parámetros no estén vacíos
    if not semana:
        messages.error(request, 'Debe seleccionar por lo menos una semana.')
        return redirect('resumen_seguimiento_list_heavens')

    # Filtrar los datos basados en los parámetros
    pedidos = Pedido.objects.all()
    if semana:
        pedidos = pedidos.filter(semana=semana)
    if exportador_id:
        pedidos = pedidos.filter(exportadora_id=exportador_id)
        try:
            exportador_nombre = Exportador.objects.get(id=exportador_id).nombre
        except Exportador.DoesNotExist:
            exportador_nombre = "Unknown"
    else:
        exportador_nombre = "Unknown"

    # Preparar los datos para el DataFrame
    data = []
    for pedido in pedidos:
        data.append({
            'Week': pedido.semana,
            'Order No.': pedido.id,
            'Exporter': pedido.exportadora.nombre if pedido.exportadora else '',
            'Customer': pedido.cliente.nombre if pedido.cliente else '',
            'Destination': pedido.destino.codigo if pedido.destino else '',
            'Products': pedido.variedades,
            'Requested Boxes': pedido.total_cajas_solicitadas,
            'Shipped Boxes': pedido.total_cajas_enviadas,
            'Requested Pallets': pedido.total_piezas_solicitadas,
            'Delivery Date': pedido.fecha_entrega.strftime('%d/%m/%Y') if pedido.fecha_entrega else '',
            'Booking Responsible': pedido.responsable_reserva.nombre if pedido.responsable_reserva else '',
            'Reserve Status': pedido.estatus_reserva,
            'Airline': pedido.aerolinea.nombre if pedido.aerolinea else '',
            'Cargo Agency': pedido.agencia_carga.nombre if pedido.agencia_carga else '',
            'Order Status': pedido.estado_pedido,
            'Document Status': pedido.estado_documentos,
            'Tracking Comments': pedido.observaciones_tracking,
        })

    df = pd.DataFrame(data)

    # Crear el archivo Excel con estilo
    wb = Workbook()
    ws = wb.active

    # Definir el color de relleno rojo suave
    fill_red_soft = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Agregar el DataFrame al archivo Excel
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Si es el encabezado
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                # Pintar la fila si el estado del pedido es 'cancelado'
                if row[df.columns.get_loc('Order Status')] == 'Cancelado':
                    cell.fill = fill_red_soft

    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=summary_week_{semana}_exporter_{exportador_nombre}.xlsx'
    wb.save(response)

    return response



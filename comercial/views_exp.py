import io
import math
import time
from collections import defaultdict
from datetime import datetime, timedelta, date
from decimal import Decimal
import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test, login_required
from django.db import transaction
from django.db.models import Q, Sum
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.decorators import method_decorator
from django.views import View
from django.views.generic import TemplateView
from django.views.generic.edit import CreateView, UpdateView
from django_tables2 import SingleTableView
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from xhtml2pdf import pisa

from .forms import SearchForm, PedidoForm, EditarPedidoForm, EliminarPedidoForm, DetallePedidoForm, \
    EliminarDetallePedidoForm, EditarPedidoExportadorForm, EditarDetallePedidoForm, EditarReferenciaForm, \
    EditarPedidoSeguimientoForm, FiltroSemanaExportadoraForm, SearchFormReferencias, EditarPedidoFormDos, \
    EditarPedidoFormCartera, EditarPedidoFormUtilidades, EditarDetallePedidoDosForm, EditarDetallePedidoTresForm, \
    ExportSearchForm, ExportSearchFormSeguimientos
from .models import Pedido, DetallePedido, Referencias, AutorizacionCancelacion, Presentacion, Exportador, Cliente
from .resources import obtener_datos_con_totales_cliente, crear_archivo_excel_cliente, \
    crear_archivo_excel_enviar_cliente, obtener_datos_con_totales_enviar_cliente
from .tables import PedidoTable, DetallePedidoTable, PedidoExportadorTable, CarteraPedidoTable, UtilidadPedidoTable, \
    ResumenPedidoTable, ReferenciasTable, SeguimienosTable, SeguimienosResumenTable


def es_miembro_del_grupo(nombre_grupo):
    def es_miembro(user):
        return user.groups.filter(name=nombre_grupo).exists()

    return es_miembro

# ---------------------------------- Funcion que exporta las utilidades a Excel ---------------------------------------

@login_required
@user_passes_test(lambda u: any(es_miembro_del_grupo(grupo)(u) for grupo in ['Heavens', 'Etnico', 'Fieldex', 'Juan_Matas', 'CI_Dorado']), login_url='home')
def exportar_utilidades_excel(request):
    # Determinar a qué grupo pertenece el usuario
    grupo = None
    if es_miembro_del_grupo('Heavens')(request.user):
        grupo = 'Heavens'
    elif es_miembro_del_grupo('Etnico')(request.user):
        grupo = 'Etnico'
    elif es_miembro_del_grupo('Fieldex')(request.user):
        grupo = 'Fieldex'
    elif es_miembro_del_grupo('Juan_Matas')(request.user):
        grupo = 'Juan_Matas'
    elif es_miembro_del_grupo('CI_Dorado')(request.user):
        grupo = 'CI_Dorado'

    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()

    # Definir estilos mejorados
    header_font = Font(bold=True, color="FFFFFF", name='Arial', size=11)
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    total_font = Font(bold=True, color="FFFFFF", name='Arial', size=11)
    total_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    total_align = Alignment(horizontal="center", vertical="center")
    
    subtotal_font = Font(bold=True, name='Arial', size=10)
    subtotal_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    date_format = 'DD-MM-YYYY'
    number_format = '$#,##0.00'
    percentage_format = '0.00%'

    # Definir el color de relleno rojo suave
    fill_red_soft = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Hoja 1: Utilidades Totales General
    worksheet1 = workbook.active
    worksheet1.title = 'Utilidades Totales General'

    # Encabezados
    columns = ['No. Pedido', 'Fecha Entrega Pedido', 'Cliente', 'Exportador', 'AWB', 'Fecha Pago Cliente', 'No Factura',
               'Valor Total Factura USD', 'Valor Pagado Cliente', 'Estado Factura', 'T Cajas Enviadas',
               'Trm Monetizacion',
               'TRM Banrep', 'Valor Utilidad USD', 'Valor Utilidad Pesos', 'Documento Cobro Utilidad',
               'Fecha Pago Utilidad', 'Diferencia O Abono', 'Estado Utilidad', 'Cobrar Utilidad']
    
    # Ajuste automático del ancho de las columnas según su contenido
    column_widths = [10, 15, 20, 15, 12, 15, 15, 15, 15, 15, 12, 12, 12, 15, 15, 20, 15, 15, 15, 10]
    
    # Aplicar encabezados y ancho de columnas
    for col_num, (column_title, width) in enumerate(zip(columns, column_widths), start=1):
        cell = worksheet1.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        worksheet1.column_dimensions[worksheet1.cell(row=1, column=col_num).column_letter].width = width

    # Crear un diccionario para almacenar los totales de Utilidades por exportadora
    totales_por_utilidad_usd = defaultdict(Decimal)
    totales_no_cobrables_por_exportadora = defaultdict(Decimal)
    totales_cobrados_por_exportadora = defaultdict(Decimal)
    totales_por_cobrar_por_exportadora = defaultdict(Decimal)
    
    # Totales generales para el resumen
    total_general_utilidad_usd = Decimal('0.00')
    total_general_no_cobrable = Decimal('0.00')
    total_general_cobrado = Decimal('0.00')
    total_general_por_cobrar = Decimal('0.00')
    total_cajas_enviadas = 0
    clientes_unicos = set()
    exportadoras_unicas = set()

    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Construir el filtro base según las fechas
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtro para las fechas
        queryset_filter = Q(fecha_entrega__gte=fecha_inicial, fecha_entrega__lte=fecha_final)
    else:
        queryset_filter = Q()  # Filtro vacío (selecciona todos)

    # Aplicar el filtro por grupo/exportadora
    if grupo == 'Heavens':
        # Si es Heavens, muestra TODOS los registros (de todas las exportadoras)
        queryset = Pedido.objects.filter(queryset_filter)
    elif grupo:
        # Para otros grupos, filtrar solo los pedidos de su exportadora
        queryset = Pedido.objects.filter(queryset_filter & Q(exportadora__nombre=grupo))
    else:
        # Si no tiene grupo específico (no debería ocurrir por el decorador)
        queryset = Pedido.objects.filter(queryset_filter)

    # Obtener los datos y acumular estadísticas
    row_num = 2  # Comenzar después del encabezado
    for pedido in queryset:
        valor_utilidad_usd = pedido.valor_total_utilidad_usd or Decimal('0.00')
        
        # Acumular estadísticas
        exportadora = pedido.exportadora.nombre
        exportadoras_unicas.add(exportadora)
        clientes_unicos.add(pedido.cliente.nombre)
        total_cajas_enviadas += pedido.total_cajas_enviadas or 0
        
        # Calcular totales por exportadora
        totales_por_utilidad_usd[exportadora] += valor_utilidad_usd
        total_general_utilidad_usd += valor_utilidad_usd
        
        # Clasificar utilidades según su estado
        if pedido.estado_utilidad == "Factura en abono" or pedido.estado_utilidad == "Pendiente Pago Cliente":
            totales_no_cobrables_por_exportadora[exportadora] += valor_utilidad_usd
            total_general_no_cobrable += valor_utilidad_usd
            
        if pedido.fecha_pago_utilidad is not None and pedido.documento_cobro_utilidad is not None:
            totales_cobrados_por_exportadora[exportadora] += valor_utilidad_usd
            total_general_cobrado += valor_utilidad_usd
            
        if pedido.fecha_pago_utilidad is None and pedido.estado_factura == "Pagada" and (
                pedido.estado_utilidad == "Por Facturar" or pedido.estado_utilidad == "Facturada"):
            totales_por_cobrar_por_exportadora[exportadora] += valor_utilidad_usd
            total_general_por_cobrar += valor_utilidad_usd

        # Determinar si se debe cobrar la utilidad
        cobrar_utilidad = "Sí" if pedido.estado_utilidad == "Por Facturar" or pedido.estado_utilidad == "Facturada" else "No"
        
        # Agregar fila de datos
        row = [
            pedido.pk,
            pedido.fecha_entrega,
            pedido.cliente.nombre,
            exportadora,
            pedido.awb,
            pedido.fecha_pago,
            pedido.numero_factura,
            pedido.valor_total_factura_usd,
            pedido.valor_pagado_cliente_usd,
            pedido.estado_factura,
            pedido.total_cajas_enviadas,
            pedido.trm_monetizacion,
            pedido.tasa_representativa_usd_diaria,
            valor_utilidad_usd,
            pedido.valor_utilidad_pesos,
            pedido.documento_cobro_utilidad,
            pedido.fecha_pago_utilidad,
            pedido.diferencia_por_abono,
            pedido.estado_utilidad,
            cobrar_utilidad,
        ]
        
        for col_num, cell_value in enumerate(row, start=1):
            cell = worksheet1.cell(row=row_num, column=col_num, value=cell_value)
            
            # Aplicar formato a fechas
            if col_num in [2, 6, 17]:  # Columnas de fechas
                if cell_value:
                    cell.number_format = date_format
            
            # Aplicar formato de moneda a las columnas específicas
            if col_num in [8, 9, 12, 13, 14, 18]:
                if cell_value:
                    cell.number_format = number_format
                    
            # Pintar la fila si el numero_factura es 'Pedido Cancelado'
            if pedido.numero_factura == 'Pedido Cancelado':
                cell.fill = fill_red_soft
        
        row_num += 1

    # Añadir totales al final de la primera hoja
    row_num += 2
    total_row = row_num
    worksheet1.cell(row=total_row, column=1, value="TOTALES GENERALES")
    worksheet1.cell(row=total_row, column=1).font = total_font
    worksheet1.cell(row=total_row, column=1).fill = total_fill
    worksheet1.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=13)
    
    # Total de Utilidades USD
    cell = worksheet1.cell(row=total_row, column=14, value=total_general_utilidad_usd)
    cell.font = total_font
    cell.fill = total_fill
    cell.number_format = number_format
    
    # Total de Utilidades en Pesos
    total_pesos = sum(p.valor_utilidad_pesos or Decimal('0.00') for p in queryset)
    cell = worksheet1.cell(row=total_row, column=15, value=total_pesos)
    cell.font = total_font
    cell.fill = total_fill
    cell.number_format = number_format

    # Hoja 2: Totales por Exportadora con formato mejorado
    worksheet2 = workbook.create_sheet(title='Resumen por Exportadora')
    
    # Título del informe - Ajustado según el grupo
    title_text = "RESUMEN DE UTILIDADES"
    if grupo and grupo != 'Heavens':
        title_text += f" PARA {grupo.upper()}"
    else:
        title_text += " POR EXPORTADORA"
        
    title_cell = worksheet2.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=14, name='Arial')
    worksheet2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Período del informe
    if fecha_inicial_str and fecha_final_str:
        period_text = f"Período: {fecha_inicial_str} al {fecha_final_str}"
        period_cell = worksheet2.cell(row=2, column=1, value=period_text)
        period_cell.font = Font(italic=True, size=10)
        worksheet2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
        period_cell.alignment = Alignment(horizontal="center")
        start_row = 4
    else:
        start_row = 3
    
    # Encabezados para la segunda hoja
    headers = ["Exportadora", "Total Utilidades USD", "Utilidades No Cobrables", "Utilidades Cobradas", "Por Cobrar"]
    header_widths = [30, 20, 25, 20, 20]
    
    for col_num, (header, width) in enumerate(zip(headers, header_widths), start=1):
        cell = worksheet2.cell(row=start_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        worksheet2.column_dimensions[worksheet2.cell(row=start_row, column=col_num).column_letter].width = width
    
    # Agregar totales a la segunda hoja
    row_num = start_row + 1
    for exportadora in sorted(totales_por_utilidad_usd.keys()):
        worksheet2.cell(row=row_num, column=1, value=exportadora)
        
        cell = worksheet2.cell(row=row_num, column=2, value=totales_por_utilidad_usd[exportadora])
        cell.number_format = number_format
        
        cell = worksheet2.cell(row=row_num, column=3, value=totales_no_cobrables_por_exportadora[exportadora])
        cell.number_format = number_format
        
        cell = worksheet2.cell(row=row_num, column=4, value=totales_cobrados_por_exportadora[exportadora])
        cell.number_format = number_format
        
        cell = worksheet2.cell(row=row_num, column=5, value=totales_por_cobrar_por_exportadora[exportadora])
        cell.number_format = number_format
        
        row_num += 1

    # Agregar totales generales
    row_num += 1
    totals_row = row_num
    
    cell = worksheet2.cell(row=totals_row, column=1, value="TOTAL GENERAL")
    cell.font = total_font
    cell.fill = total_fill
    
    cell = worksheet2.cell(row=totals_row, column=2, value=total_general_utilidad_usd)
    cell.font = total_font
    cell.fill = total_fill
    cell.number_format = number_format
    
    cell = worksheet2.cell(row=totals_row, column=3, value=total_general_no_cobrable)
    cell.font = total_font
    cell.fill = total_fill
    cell.number_format = number_format
    
    cell = worksheet2.cell(row=totals_row, column=4, value=total_general_cobrado)
    cell.font = total_font
    cell.fill = total_fill
    cell.number_format = number_format
    
    cell = worksheet2.cell(row=totals_row, column=5, value=total_general_por_cobrar)
    cell.font = total_font
    cell.fill = total_fill
    cell.number_format = number_format
    
    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                           content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    # Nombre del archivo con fecha si se especificó un rango
    filename_parts = []
    if grupo and grupo != 'Heavens':
        filename_parts.append(f"{grupo}")
    filename_parts.append("utilidades_pedidos")
    
    if fecha_inicial_str and fecha_final_str:
        filename_parts.append(f"{fecha_inicial_str}_a_{fecha_final_str}")
    
    filename = "_".join(filename_parts) + ".xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


# ---------------------------------- Funcion que exporta las utilidades a Excel ---------------------------------------

# ---------------------------------- Funcion que exporta los pedidos a Excel ------------------------------------------

@login_required
@user_passes_test(lambda u: any(es_miembro_del_grupo(grupo)(u) for grupo in ['Heavens', 'Etnico', 'Fieldex', 'Juan_Matas', 'CI_Dorado']), login_url='home')
def exportar_pedidos_excel_general(request):
    # Determinar a qué grupo pertenece el usuario
    grupo = None
    if es_miembro_del_grupo('Heavens')(request.user):
        grupo = 'Heavens'
    elif es_miembro_del_grupo('Etnico')(request.user):
        grupo = 'Etnico'
    elif es_miembro_del_grupo('Fieldex')(request.user):
        grupo = 'Fieldex'
    elif es_miembro_del_grupo('Juan_Matas')(request.user):
        grupo = 'Juan_Matas'
    elif es_miembro_del_grupo('CI_Dorado')(request.user):
        grupo = 'CI_Dorado'
        
    # 1. Crear un buffer en memoria para almacenar el archivo Excel
    output = io.BytesIO()

    # 2. Crear el Workbook en modo write_only para optimizar el uso de memoria
    workbook = Workbook(write_only=True)
    ws = workbook.create_sheet(title='Pedidos')

    # 3. Definir estilos para los encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e0c42", end_color="1e0c42", fill_type="solid")
    sub_header_fill = PatternFill(start_color="0B6FA4", end_color="0B6FA4", fill_type="solid")

    # 4. Definir los encabezados para Pedido y DetallePedido
    pedido_headers = [
        'No', 'Cliente', 'Semana', 'Fecha Solicitud', 'Fecha Entrega', 'Fecha Llegada',
        'Exportador', 'Subexportadora', 'Intermediario', 'Dias Cartera', 'AWB',
        'Destino', 'Aerolinea', 'Agencia De Carga', 'Responsable Reserva', 'Numero Factura',
        'Total Cajas Solicitadas', 'Total Cajas Enviadas', 'Peso Bruto Solicitado',
        'Peso Bruto Enviado', 'Pallets Solicitados', 'Pallets Enviados', 'Peso AWB',
        'ETA', 'ETD', 'Variedades', 'Descuento Comercial', 'No NC', 'Motivo NC',
        'Valor Total NC', 'Valor Pagado Cliente', 'Estado Factura', 'Utilidad Bancaria USD',
        'Fecha Pago Cliente', 'TRM Monetización', 'Fecha Monetización', 'Trm Banrep',
        'Trm Cotización', 'Diferencia Pago', 'Dias Vencimiento', 'Valor Total Factura USD',
        'Valor Utilidad USD', 'Valor Utilidad Pesos', 'Documento Cobro Utilidad',
        'Fecha Pago Utilidad', 'Estado Utilidad', 'Estado Cancelacion', 'Estado Documentos',
        'Estado Reserva', 'Termo', 'Diferencia AWB/Factura', 'Eta Real', 'Estado Pedido',
        'Observaciones Tracking', 'Observaciones Generales'
    ]

    detalle_headers = [
        'Pedido', 'F Entrega', 'Exportador', 'Cliente', 'Fruta', 'Presentacion', 'Cajas Solicitadas',
        'Peso Presentacion', 'kilos', 'Cajas Enviadas', 'Kilos Enviados', 'Diferencia', 'Tipo Caja',
        'Referencia', 'Stiker', 'Lleva Contenedor', 'Ref Contenedor', 'Cant Contenedor', 'Tarifa utilidad', 'Tarifa Recuperacion',
        'Valor x Caja USD', 'Valor X Producto', 'No Cajas NC', 'Valor NC', 'Afecta utilidad',
        'Valor Total utilidad Producto', 'Precio Proforma', 'Observaciones'
    ]

    # 5. Verificar si el usuario incluyó detalles
    incluir_detalles = request.POST.get('incluir_detalles') == 'true'

    # 6. Obtener filtros de fecha desde el formulario (si existen)
    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # 7. Filtrar los pedidos según fechas
    try:
        if fecha_inicial_str and fecha_final_str:
            fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
            fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')
            base_filter = Q(fecha_entrega__gte=fecha_inicial, fecha_entrega__lte=fecha_final)
        else:
            base_filter = Q()
            
        # Filtrar según el grupo del usuario
        if grupo == 'Heavens':
            # Si es Heavens, muestra TODOS los registros (de todas las exportadoras)
            pedidos_qs = Pedido.objects.filter(base_filter)
        elif grupo:
            # Para otros grupos, filtrar solo los pedidos de su exportadora
            pedidos_qs = Pedido.objects.filter(base_filter & Q(exportadora__nombre=grupo))
        else:
            # Si no tiene grupo específico (no debería ocurrir por el decorador)
            pedidos_qs = Pedido.objects.filter(base_filter)
            
        pedidos_qs = pedidos_qs.select_related(
            'cliente',
            'exportadora',
            'subexportadora',
            'intermediario',
            'destino',
            'aerolinea',
            'agencia_carga',
            'responsable_reserva'
        ).iterator(chunk_size=50)
            
    except ValueError:
        return HttpResponse("Fecha inválida", status=400)

    for pedido in pedidos_qs:
        # 7.1 Escribir una fila de separación para mayor legibilidad (opcional)
        ws.append([])

        # 7.2 Escribir una fila de cabecera para "Pedido"
        row_header_pedido = [WriteOnlyCell(ws, value='PEDIDO:')]
        row_header_pedido[0].font = Font(bold=True, color='FFFFFF')
        row_header_pedido[0].fill = PatternFill(start_color="0B6FA4", end_color="0B6FA4", fill_type="solid")
        ws.append(row_header_pedido)

        # 7.3 Escribir los encabezados de Pedido
        pedido_header_cells = []
        for titulo in pedido_headers:
            celda = WriteOnlyCell(ws, value=titulo)
            celda.font = header_font
            celda.fill = header_fill
            pedido_header_cells.append(celda)
        ws.append(pedido_header_cells)

        # 7.4 Escribir los datos del Pedido
        pedido_data = [
            pedido.pk,
            pedido.cliente.nombre if pedido.cliente else '',
            pedido.semana,
            pedido.fecha_solicitud.strftime('%Y-%m-%d') if pedido.fecha_solicitud else '',
            pedido.fecha_entrega.strftime('%Y-%m-%d') if pedido.fecha_entrega else '',
            pedido.fecha_llegada.strftime('%Y-%m-%d') if pedido.fecha_llegada else '',
            pedido.exportadora.nombre if pedido.exportadora else '',
            pedido.subexportadora.nombre if pedido.subexportadora else '',
            pedido.intermediario.nombre if pedido.intermediario else '',
            pedido.dias_cartera,
            pedido.awb,
            pedido.destino.codigo if pedido.destino else '',
            pedido.aerolinea.nombre if pedido.aerolinea else '',
            pedido.agencia_carga.nombre if pedido.agencia_carga else '',
            pedido.responsable_reserva.nombre if pedido.responsable_reserva else '',
            pedido.numero_factura,
            pedido.total_cajas_solicitadas,
            pedido.total_cajas_enviadas,
            pedido.total_peso_bruto_solicitado,
            pedido.total_peso_bruto_enviado,
            pedido.total_piezas_solicitadas,
            pedido.total_piezas_enviadas,
            pedido.peso_awb,
            pedido.eta.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S') if pedido.eta else '',
            pedido.etd.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S') if pedido.etd else '',
            pedido.variedades,
            pedido.descuento,
            pedido.nota_credito_no,
            pedido.motivo_nota_credito,
            pedido.valor_total_nota_credito_usd,
            pedido.valor_pagado_cliente_usd,
            pedido.estado_factura,
            pedido.utilidad_bancaria_usd,
            pedido.fecha_pago.strftime('%Y-%m-%d') if pedido.fecha_pago else '',
            pedido.trm_monetizacion,
            pedido.fecha_monetizacion.strftime('%Y-%m-%d') if pedido.fecha_monetizacion else '',
            pedido.tasa_representativa_usd_diaria,
            pedido.trm_cotizacion,
            pedido.diferencia_por_abono,
            pedido.dias_de_vencimiento,
            pedido.valor_total_factura_usd,
            pedido.valor_total_utilidad_usd,
            pedido.valor_utilidad_pesos,
            pedido.documento_cobro_utilidad,
            pedido.fecha_pago_utilidad.strftime('%Y-%m-%d') if pedido.fecha_pago_utilidad else '',
            pedido.estado_utilidad,
            pedido.estado_cancelacion,
            pedido.estado_documentos,
            pedido.estatus_reserva,
            pedido.termo,
            pedido.diferencia_peso_factura_awb,
            pedido.eta_real.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S') if pedido.eta_real else '',
            pedido.estado_pedido,
            pedido.observaciones_tracking,
            pedido.observaciones
        ]
        ws.append(pedido_data)

        if incluir_detalles:
            # Solamente si se marca 'incluir_detalles'
            row_header_detalle = [WriteOnlyCell(ws, value='DETALLES DEL PEDIDO:')]
            row_header_detalle[0].font = Font(bold=True, color='FFFFFF')
            row_header_detalle[0].fill = PatternFill(start_color="0B6FA4", end_color="0B6FA4", fill_type="solid")
            ws.append(row_header_detalle)

            detalle_header_cells = []
            for titulo in detalle_headers:
                celda = WriteOnlyCell(ws, value=titulo)
                celda.font = header_font
                celda.fill = header_fill
                detalle_header_cells.append(celda)
            ws.append(detalle_header_cells)

            detalles_qs = DetallePedido.objects.filter(pedido=pedido).select_related(
                'pedido__exportadora',
                'pedido__cliente',
                'fruta',
                'presentacion',
                'tipo_caja',
                'referencia'
            ).iterator(chunk_size=50)

            for detalle in detalles_qs:
                detalle_row = [
                    detalle.pedido.pk,
                    detalle.pedido.fecha_entrega.strftime('%Y-%m-%d') if detalle.pedido.fecha_entrega else '',
                    detalle.pedido.exportadora.nombre if detalle.pedido.exportadora else '',
                    detalle.pedido.cliente.nombre if detalle.pedido.cliente else '',
                    detalle.fruta.nombre if detalle.fruta else '',
                    detalle.presentacion.nombre if detalle.presentacion else '',
                    detalle.cajas_solicitadas,
                    detalle.presentacion_peso,
                    detalle.kilos,
                    detalle.cajas_enviadas,
                    detalle.kilos_enviados,
                    detalle.diferencia,
                    detalle.tipo_caja.nombre if detalle.tipo_caja else '',
                    detalle.referencia.nombre if detalle.referencia else '',
                    detalle.stickers,
                    detalle.lleva_contenedor,
                    detalle.referencia_contenedor,
                    detalle.cantidad_contenedores,
                    detalle.tarifa_utilidad,
                    detalle.tarifa_recuperacion,
                    detalle.valor_x_caja_usd,
                    detalle.valor_x_producto,
                    detalle.no_cajas_nc,
                    detalle.valor_nota_credito_usd,
                    detalle.afecta_utilidad,
                    detalle.valor_total_utilidad_x_producto,
                    detalle.precio_proforma,
                    detalle.observaciones,
                ]
                ws.append(detalle_row)
            # =============== FIN SECCIÓN DETALLES (CONDICIONAL) ===============
    # 9. Guardar y retornar el archivo
    workbook.save(output)
    output.seek(0)
    
    # Personalizar el nombre del archivo según el grupo
    filename_parts = []
    if grupo and grupo != 'Heavens':
        filename_parts.append(f"{grupo}")
    filename_parts.append("pedidos")
    
    if fecha_inicial_str and fecha_final_str:
        filename_parts.append(f"{fecha_inicial_str}_a_{fecha_final_str}")
    
    filename = "_".join(filename_parts) + ".xlsx"
    
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response





